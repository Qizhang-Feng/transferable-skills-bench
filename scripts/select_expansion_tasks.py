#!/usr/bin/env python3
"""
Step 1: Select expansion tasks from the 60 audited tasks.

Reads task-audit-60-v1.json, excludes pilot tasks, applies filters,
and for each candidate extracts xlsx schema info (sheets, headers,
formula count, merged cells) to support perturbation spec generation.
"""

import json
import os
import sys
import openpyxl
from collections import OrderedDict

AUDIT_PATH = "scratch/task-audit-60-v1.json"
PILOT_SPLIT_PATH = "scratch/pilot-task-split.json"
VERIFIED_DIR = "vendor/SpreadsheetBench/data/spreadsheetbench_verified_400"
DATASET_PATH = os.path.join(VERIFIED_DIR, "dataset.json")
OUTPUT_PATH = "scratch/expansion-task-candidates.json"


def get_pilot_task_ids():
    """Get all task IDs used in the pilot (including excluded ones)."""
    with open(PILOT_SPLIT_PATH) as f:
        pilot = json.load(f)
    ids = set()
    for fam in pilot.values():
        for t in fam["tasks"]:
            ids.add(str(t))
    return ids


def extract_xlsx_schema(task_id):
    """Extract schema info from the task's init.xlsx."""
    spread_dir = os.path.join(VERIFIED_DIR, "spreadsheet", str(task_id))
    init_file = os.path.join(spread_dir, f"1_{task_id}_init.xlsx")
    
    if not os.path.exists(init_file):
        return {"error": f"File not found: {init_file}"}
    
    try:
        wb = openpyxl.load_workbook(init_file, data_only=True)
    except Exception as e:
        return {"error": f"Cannot open: {e}"}
    
    schema = {"sheets": []}
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet_info = {
            "name": ws_name,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_cells": len(ws.merged_cells.ranges),
        }
        
        # Count formulas
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    formula_count += 1
        sheet_info["formula_count"] = formula_count
        
        # Extract headers (first non-empty row)
        headers = []
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(max_row=min(5, ws.max_row or 1)), 1):
            row_vals = [c.value for c in row if c.value is not None]
            if len(row_vals) >= 2:  # at least 2 non-empty cells = likely header
                header_row = row_idx
                headers = [c.value for c in row]
                break
        
        sheet_info["header_row"] = header_row
        sheet_info["headers"] = [str(h) if h else None for h in headers]
        
        schema["sheets"].append(sheet_info)
    
    wb.close()
    return schema


def get_task_instruction(task_id, dataset):
    """Get full instruction from dataset.json."""
    for d in dataset:
        if str(d["id"]) == str(task_id):
            return d.get("instruction", ""), d.get("answer_position", "")
    return "", ""


def check_instruction_mentions(instruction, schema):
    """Check which column names and sheet names from the xlsx appear in the instruction."""
    instruction_lower = instruction.lower()
    
    mentioned_columns = []
    mentioned_sheets = []
    
    for sheet in schema.get("sheets", []):
        # Check sheet name
        sname = sheet["name"]
        if sname.lower() in instruction_lower and len(sname) > 1:
            mentioned_sheets.append(sname)
        
        # Check headers
        for h in sheet.get("headers", []):
            if h and len(str(h)) > 1 and str(h).lower() in instruction_lower:
                mentioned_columns.append(str(h))
    
    return list(set(mentioned_columns)), list(set(mentioned_sheets))



def main():
    # Load data
    with open(AUDIT_PATH) as f:
        audit = json.load(f)
    with open(DATASET_PATH) as f:
        dataset = json.load(f)
    
    pilot_ids = get_pilot_task_ids()
    all_tasks = audit["tasks"]
    
    print(f"Total audited tasks: {len(all_tasks)}")
    print(f"Pilot task IDs: {sorted(pilot_ids)}")
    
    # Separate into pilot vs remaining
    remaining = [t for t in all_tasks if str(t["task_id"]) not in pilot_ids]
    print(f"Remaining after excluding pilot: {len(remaining)}")
    
    # Apply filters
    candidates = []
    excluded = []
    
    for task in remaining:
        tid = str(task["task_id"])
        perturbability = task.get("perturbability", "unknown")
        family = task.get("workflow_family_v1", "unknown")
        
        # Filter 1: perturbability must be easy or medium
        if perturbability == "hard":
            excluded.append({"task_id": tid, "reason": "perturbability=hard"})
            continue
        
        # Get xlsx schema
        schema = extract_xlsx_schema(tid)
        if "error" in schema:
            excluded.append({"task_id": tid, "reason": f"schema error: {schema['error']}"})
            continue
        
        # Filter 2: check for merged cells (total across all sheets)
        total_merged = sum(s.get("merged_cells", 0) for s in schema.get("sheets", []))
        
        # Filter 3: check total formula count
        total_formulas = sum(s.get("formula_count", 0) for s in schema.get("sheets", []))
        
        # Filter 4: avoid single-letter sheet names
        single_letter_sheets = [s["name"] for s in schema.get("sheets", []) 
                                if len(s["name"]) == 1]
        
        # Get instruction and check mentions
        instruction, answer_pos = get_task_instruction(tid, dataset)
        mentioned_cols, mentioned_sheets = check_instruction_mentions(instruction, schema)
        
        # Filter 5: instruction must mention at least one column or sheet name
        has_mentions = len(mentioned_cols) > 0 or len(mentioned_sheets) > 0
        
        # Build candidate record
        candidate = {
            "task_id": tid,
            "workflow_family_v1": family,
            "workflow_family_v1_secondary": task.get("workflow_family_v1_secondary"),
            "perturbability": perturbability,
            "instruction_type": task.get("instruction_type"),
            "answer_position": task.get("answer_position"),
            "instruction_preview": task.get("instruction_preview", "")[:150],
            "schema": {
                "n_sheets": len(schema.get("sheets", [])),
                "sheet_names": [s["name"] for s in schema.get("sheets", [])],
                "total_formulas": total_formulas,
                "total_merged_cells": total_merged,
                "single_letter_sheets": single_letter_sheets,
            },
            "instruction_mentions": {
                "columns": mentioned_cols,
                "sheets": mentioned_sheets,
            },
            "filters": {
                "perturbability_ok": perturbability in ("easy", "medium"),
                "has_instruction_mentions": has_mentions,
                "merged_cells_low": total_merged <= 5,
                "formulas_manageable": total_formulas <= 50,
                "no_single_letter_sheets": len(single_letter_sheets) == 0,
            },
        }
        
        # Compute overall eligibility
        filters = candidate["filters"]
        all_pass = all(filters.values())
        soft_pass = (filters["perturbability_ok"] and filters["has_instruction_mentions"])
        
        candidate["eligible"] = all_pass
        candidate["eligible_soft"] = soft_pass  # passes hard requirements, may have warnings
        
        if not filters["has_instruction_mentions"]:
            excluded.append({"task_id": tid, "reason": "no column/sheet names in instruction"})
        else:
            candidates.append(candidate)
    
    # Sort by eligibility then family
    candidates.sort(key=lambda c: (not c["eligible"], c["workflow_family_v1"], c["task_id"]))
    
    # Summary
    eligible_strict = [c for c in candidates if c["eligible"]]
    eligible_soft = [c for c in candidates if c["eligible_soft"]]
    
    print(f"\n{'='*60}")
    print(f"RESULTS")
    print(f"{'='*60}")
    print(f"Candidates with instruction mentions: {len(candidates)}")
    print(f"  Eligible (all filters pass): {len(eligible_strict)}")
    print(f"  Eligible (soft - has mentions): {len(eligible_soft)}")
    print(f"Excluded: {len(excluded)}")
    
    # Family distribution
    print(f"\nFamily distribution (eligible strict):")
    from collections import Counter
    fam_counts = Counter(c["workflow_family_v1"] for c in eligible_strict)
    for fam, count in sorted(fam_counts.items()):
        print(f"  {fam}: {count}")
    
    print(f"\nFamily distribution (all candidates):")
    fam_counts_all = Counter(c["workflow_family_v1"] for c in candidates)
    for fam, count in sorted(fam_counts_all.items()):
        print(f"  {fam}: {count}")
    
    # Print excluded
    print(f"\nExcluded tasks:")
    for e in excluded:
        print(f"  {e['task_id']}: {e['reason']}")
    
    # Print candidates with warnings
    print(f"\nCandidates with warnings:")
    for c in candidates:
        if c["eligible_soft"] and not c["eligible"]:
            warnings = [k for k, v in c["filters"].items() if not v]
            print(f"  {c['task_id']} ({c['workflow_family_v1']}): {', '.join(warnings)}")
    
    # Save output
    output = {
        "metadata": {
            "total_audited": len(all_tasks),
            "pilot_ids": sorted(pilot_ids),
            "remaining": len(remaining),
            "candidates": len(candidates),
            "eligible_strict": len(eligible_strict),
            "eligible_soft": len(eligible_soft),
            "excluded": len(excluded),
        },
        "candidates": candidates,
        "excluded": excluded,
    }
    
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    print(f"\nSaved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
