#!/usr/bin/env python3
"""
Run no_skill baseline across 4 conditions to diagnose perturbation neutrality.
12 tasks × 4 conditions × 4 repeats = 192 runs, all in one batch.
"""

import json, os, sys, tempfile, shutil, subprocess, re, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

import openpyxl
import pandas as pd

sys.path.insert(0, "vendor/PowerComputeCustomImage/configuration/rewards/Inference/APIInference")

VERIFIED_DIR = "vendor/SpreadsheetBench/data/spreadsheetbench_verified_400"
EXCLUDED_TASKS = {"48982", "38703", "304-35", "333-29"}

CONDITION_DIRS = {
    "original": None,  # use VERIFIED_DIR directly
    "perturbed": "scratch/pilot_perturbed",
    "rename_only": "scratch/pilot_perturbed_rename_only",
    "layout_only": "scratch/pilot_perturbed_layout_only",
}

# ---- LLM ----
_claude = None
def call_llm(prompt):
    global _claude
    if _claude is None:
        from api_invocation import ClaudeInference
        _claude = ClaudeInference()
    return _claude.inference(
        prompt_dict={"system": "You are a spreadsheet expert.", "user": prompt},
        retry=3, guardrails=[],
        model_id="us.anthropic.claude-sonnet-4-5-20250929-v1:0",
        max_tokens=4096, temperature=0,
    )

# ---- Code extraction (robust) ----
def extract_code(response):
    if "```python" in response:
        return response.split("```python")[1].split("```")[0].strip()
    if "```" in response:
        blocks = []
        parts = response.split("```")
        for i in range(1, len(parts), 2):
            block = parts[i]
            lines = block.split('\n')
            if lines and lines[0].strip().lower() in ('python', 'py', 'python3', ''):
                block = '\n'.join(lines[1:])
            blocks.append(block.strip())
        if blocks:
            return max(blocks, key=len)
    lines = response.split('\n')
    code_lines = []
    in_code = False
    for line in lines:
        s = line.strip()
        if not in_code and s.startswith(('import ', 'from ', 'def ', 'class ', '# ', 'wb ', 'ws ')):
            in_code = True
        if in_code:
            code_lines.append(line)
    if code_lines:
        return '\n'.join(code_lines).strip()
    return response.strip()

# ---- Evaluation ----
def transform_value(v):
    if isinstance(v, (int, float)): return round(float(v), 2)
    elif isinstance(v, datetime.time): return str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        d = v - datetime.datetime(1899, 12, 30)
        return round(d.days + d.seconds / 86400.0, 0)
    elif isinstance(v, str):
        try: return round(float(v), 2)
        except: pass
    return v

def compare_cell(v1, v2):
    v1, v2 = transform_value(v1), transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""): return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None): return True
    if type(v1) != type(v2): return False
    return v1 == v2

def evaluate(golden_path, output_path, answer_position):
    if not os.path.exists(output_path): return False
    try:
        wb_gt = openpyxl.load_workbook(golden_path, data_only=True)
        wb_out = openpyxl.load_workbook(output_path, data_only=True)
    except: return False
    
    ap = re.sub(r"'([^'!]+)!'", r"'\1'!", answer_position)
    ap = ap.rstrip("'").rstrip('"')
    parts = re.split(r",(?='|[A-Z])", ap)
    
    for part in parts:
        part = part.strip().strip("'").strip('"')
        if '!' in part:
            sp, cp = part.rsplit('!', 1)
            sheet = sp.strip("'")
        else:
            sheet = wb_gt.sheetnames[0]
            cp = part
        cp = cp.strip("'")
        if sheet not in wb_out.sheetnames: return False
        ws_gt, ws_out = wb_gt[sheet], wb_out[sheet]
        
        if ':' in cp:
            start, end = cp.split(':')
            sc, sr = '', ''
            for c in start:
                if c.isdigit(): sr += c
                else: sc += c
            ec, er = '', ''
            for c in end:
                if c.isdigit(): er += c
                else: ec += c
            def c2n(name):
                n = 0
                for c in name: n = n*26 + ord(c)-64
                return n
            def n2c(n):
                s = ''
                while n > 0:
                    n, r = divmod(n-1, 26)
                    s = chr(65+r) + s
                return s
            cells = [f"{n2c(i)}{r}" for i in range(c2n(sc), c2n(ec)+1) for r in range(int(sr), int(er)+1)]
        else:
            cells = [cp]
        
        for cn in cells:
            if not compare_cell(ws_gt[cn].value, ws_out[cn].value):
                return False
    return True

# ---- Prompt ----
def build_prompt(meta, input_path, output_path):
    try:
        ef = pd.ExcelFile(input_path)
        content = ""
        for sn in ef.sheet_names:
            df = ef.parse(sn)
            content += f"Sheet Name: {sn}\n{df.head(5).to_string()}\n{'-'*50}\n"
    except:
        content = "Error reading spreadsheet"
    
    return f"""You need to solve the given spreadsheet manipulation question.

### instruction
{meta['instruction']}

### spreadsheet_path
{input_path}

### spreadsheet_content
{content}

### instruction_type
{meta['instruction_type']}

### answer_position
{meta['answer_position']}

### output_path
{output_path}

You should generate Python code for the final solution of the question."""

# ---- Main ----
def run_one(task_id, condition, repeat):
    if condition == "original":
        spread_dir = os.path.join(VERIFIED_DIR, "spreadsheet", str(task_id))
        with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
            dataset = json.load(f)
        meta = next(d for d in dataset if str(d["id"]) == str(task_id))
    else:
        spread_dir = os.path.join(CONDITION_DIRS[condition], str(task_id))
        with open(os.path.join(spread_dir, "task_meta.json")) as f:
            meta = json.load(f)
    
    input_file = f"1_{task_id}_init.xlsx"
    golden_file = f"1_{task_id}_golden.xlsx"
    input_path = os.path.join(spread_dir, input_file)
    golden_path = os.path.join(spread_dir, golden_file)
    
    if not os.path.exists(input_path):
        return {"task_id": task_id, "condition": condition, "repeat": repeat, "pass": False, "error": "no input"}
    
    with tempfile.TemporaryDirectory() as wd:
        shutil.copy2(input_path, os.path.join(wd, input_file))
        output_path = os.path.join(wd, f"1_{task_id}_output.xlsx")
        
        try:
            response = call_llm(build_prompt(meta, os.path.join(wd, input_file), output_path))
        except Exception as e:
            return {"task_id": task_id, "condition": condition, "repeat": repeat, "pass": False, "error": str(e)[:100]}
        
        code = extract_code(response)
        code_path = os.path.join(wd, "solution.py")
        with open(code_path, 'w') as f:
            f.write(code)
        
        try:
            result = subprocess.run([sys.executable, code_path], capture_output=True, text=True, timeout=120, cwd=wd)
            executed = result.returncode == 0
        except:
            executed = False
        
        passed = False
        if executed and os.path.exists(output_path):
            passed = evaluate(golden_path, output_path, meta["answer_position"])
        
        return {"task_id": task_id, "condition": condition, "repeat": repeat, "pass": passed, "executed": executed}

if __name__ == "__main__":
    with open("scratch/pilot-task-split.json") as f:
        pilot = json.load(f)
    
    all_tasks = [str(t) for t in pilot["family_A"]["tasks"] + pilot["family_B"]["tasks"] if str(t) not in EXCLUDED_TASKS]
    conditions = ["original", "perturbed", "rename_only", "layout_only"]
    
    jobs = [(tid, cond, rep) for tid in all_tasks for cond in conditions for rep in range(4)]
    print(f"Running {len(jobs)} jobs (no_skill only, 4 conditions, 4 repeats)")
    
    results = []
    lock = threading.Lock()
    done = [0]
    
    def run_job(job):
        r = run_one(*job)
        with lock:
            done[0] += 1
            s = "✅" if r["pass"] else "❌"
            print(f"[{done[0]}/{len(jobs)}] {s} {job[0]}|{job[1]}|r{job[2]}")
        return r
    
    with ThreadPoolExecutor(max_workers=32) as ex:
        futures = {ex.submit(run_job, j): j for j in jobs}
        for f in as_completed(futures):
            results.append(f.result())
    
    out_path = "scratch/experiment_results/noskill_baseline_4cond.json"
    with open(out_path, 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"\n{'='*50}\nSUMMARY\n{'='*50}")
    for cond in conditions:
        subset = [r for r in results if r["condition"] == cond]
        p = sum(r["pass"] for r in subset)
        print(f"  {cond:15s} | {p}/{len(subset)} = {p/len(subset):.1%}")
