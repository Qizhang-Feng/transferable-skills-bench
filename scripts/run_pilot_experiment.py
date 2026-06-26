#!/usr/bin/env python3
"""
Pilot experiment runner.
Runs 12 tasks × 2 conditions × 4 artifacts = 96 experiments.
"""

import json
import os
import sys
import re
import subprocess
import tempfile
import shutil
import argparse
from pathlib import Path

import openpyxl
import pandas as pd


# ============================================================
# Config
# ============================================================

VERIFIED_DIR = "vendor/SpreadsheetBench/data/spreadsheetbench_verified_400"
PERTURBED_DIR = "scratch/pilot_perturbed_rename_only"
EXPANSION_PERTURBED_DIR = "scratch/expansion_perturbed"
ARTIFACT_DIR = "scratch/artifact_baselines"
RESULTS_DIR = "scratch/experiment_results"

ARTIFACT_TYPES = ["no_skill", "gen_skill", "local_patch"]
CONDITIONS = ["original", "perturbed"]

EXCLUDED_TASKS = {"48982", "38703", "304-35", "333-29"}

# Pilot tasks (use pilot perturbed dir)
PILOT_TASKS = {"398-14", "48365", "50324", "379-36", "52807", "32337", "46167",
               "130-9", "170-13", "44266", "177-6", "13-1"}

# Expansion tasks (use expansion perturbed dir)
EXPANSION_TASKS = {"24-23", "341-40", "1818", "31915", "36191", "49857", "58723",
                   "57558", "52575", "203-15", "395-36", "367-23", "493-5",
                   "488-14", "374-31", "486-17", "577-40"}


# ============================================================
# Prompt construction
# ============================================================

SYSTEM_PROMPT = """You are a spreadsheet expert who can manipulate spreadsheets through Python code.

{artifact_section}
You need to solve the given spreadsheet manipulation question, which contains six types of information:
- instruction: The question about spreadsheet manipulation.
- spreadsheet_path: The path of the spreadsheet file you need to manipulate.
- spreadsheet_content: The first few rows of the content of spreadsheet file.
- instruction_type: There are two values (Cell-Level Manipulation, Sheet-Level Manipulation) used to indicate whether the answer to this question applies only to specific cells or to the entire worksheet.
- answer_position: The position need to be modified or filled.
- output_path: You need to generate the modified spreadsheet file in this new path.

Below is the spreadsheet manipulation question you need to solve:
### instruction
{instruction}

### spreadsheet_path
{spreadsheet_path}

### spreadsheet_content
{spreadsheet_content}

### instruction_type
{instruction_type}

### answer_position
{answer_position}

### output_path
{output_path}

You should generate Python code for the final solution of the question.
"""


def load_artifact(artifact_type: str, family: str, task_id: str = None) -> str:
    """Load artifact content. Returns empty string for no_skill."""
    if artifact_type == "no_skill":
        return ""

    # gen_skill is family-level
    if artifact_type == "gen_skill":
        family_dir_map = {
            "formula_writeback": "family_a_formula_writeback",
            "lookup_match_merge": "family_b_lookup_match_merge",
            "conditional_partition": "family_c_conditional_partition",
            "grouped_aggregation": "family_d_grouped_aggregation",
            "structural_edit": "family_e_structural_edit",
        }
        family_dir = family_dir_map[family]
        path = os.path.join(ARTIFACT_DIR, family_dir, "generalized_skill.md")
        with open(path) as f:
            return f.read()

    # local_patch is per-task
    type_map = {
        "local_patch": "local_patch.md",
    }
    filename = type_map.get(artifact_type)
    if not filename:
        return ""
    
    # Try per-task first
    per_task_path = os.path.join(ARTIFACT_DIR, "per_task", str(task_id), filename)
    if os.path.exists(per_task_path):
        with open(per_task_path) as f:
            return f.read()
    
    # Fallback to family-level
    family_dir_map = {
        "formula_writeback": "family_a_formula_writeback",
        "lookup_match_merge": "family_b_lookup_match_merge",
        "conditional_partition": "family_c_conditional_partition",
        "grouped_aggregation": "family_d_grouped_aggregation",
        "structural_edit": "family_e_structural_edit",
    }
    family_dir = family_dir_map[family]
    path = os.path.join(ARTIFACT_DIR, family_dir, filename)
    if os.path.exists(path):
        with open(path) as f:
            return f.read()
    return ""


def gen_spreadsheet_content(input_path: str, max_rows: int = 5) -> str:
    """Read first N rows of spreadsheet as string using openpyxl.
    This avoids pandas header detection issues with title rows."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(input_path, data_only=True)
        result = ""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result += f"Sheet Name: {sheet_name}\n"
            for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows + 1, values_only=True), 1):
                vals = [str(v) if v is not None else "" for v in row]
                result += f"  Row {row_idx}: {vals}\n"
            result += "-" * 50 + "\n"
        wb.close()
        return result
    except Exception as e:
        return f"Error reading spreadsheet: {e}"


def build_prompt(task_meta: dict, input_path: str, output_path: str,
                 artifact_type: str, family: str, task_id: str = None) -> str:
    """Build the full prompt for one experiment run."""
    artifact_content = load_artifact(artifact_type, family, task_id)

    if artifact_content:
        artifact_section = (
            "### knowledge_artifact\n"
            "The following knowledge artifact may help you solve this task. "
            "Use it as reference if relevant.\n\n"
            f"{artifact_content}\n\n"
        )
    else:
        artifact_section = ""

    spreadsheet_content = gen_spreadsheet_content(input_path)

    return SYSTEM_PROMPT.format(
        artifact_section=artifact_section,
        instruction=task_meta["instruction"],
        spreadsheet_path=input_path,
        spreadsheet_content=spreadsheet_content,
        instruction_type=task_meta["instruction_type"],
        answer_position=task_meta["answer_position"],
        output_path=output_path,
    )


# ============================================================
# LLM call
# ============================================================

def call_llm(prompt: str, model: str, api_key: str, base_url: str = None) -> str:
    """Call LLM API and return response text. Supports OpenAI and Bedrock (via cross-account)."""
    if model.startswith("anthropic.") or model.startswith("us.anthropic."):
        # Use ClaudeInference from PowerComputeCustomImage for auto-refreshing cross-account access
        sys.path.insert(0, "vendor/PowerComputeCustomImage/configuration/rewards/Inference/APIInference")
        from api_invocation import ClaudeInference
        
        if not hasattr(call_llm, '_claude'):
            call_llm._claude = ClaudeInference()
        
        prompt_dict = {
            "system": "You are a spreadsheet expert who can manipulate spreadsheets through Python code.",
            "user": prompt,
        }
        return call_llm._claude.inference(
            prompt_dict=prompt_dict,
            retry=3,
            guardrails=[],
            model_id=model,
            max_tokens=4096,
            temperature=0,
        )
    else:
        # OpenAI-compatible
        from openai import OpenAI
        client = OpenAI(api_key=api_key, base_url=base_url)
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
        )
        return response.choices[0].message.content


# ============================================================
# Code execution
# ============================================================

def extract_code(response: str) -> str:
    """Extract Python code block from LLM response. Handles multiple formats."""
    # Try ```python block first
    if "```python" in response:
        code = response.split("```python")[1].split("```")[0]
        return code.strip()
    
    # Try generic ``` block
    if "```" in response:
        # Find all code blocks, pick the longest one (likely the actual solution)
        blocks = []
        parts = response.split("```")
        for i in range(1, len(parts), 2):
            block = parts[i]
            # Skip language tags on first line
            lines = block.split('\n')
            if lines and lines[0].strip().lower() in ('python', 'py', 'python3', ''):
                block = '\n'.join(lines[1:])
            blocks.append(block.strip())
        if blocks:
            # Return the longest block
            return max(blocks, key=len)
    
    # No code blocks found - try to extract Python code from mixed text
    # Look for lines that start with import/from/def/class or common Python patterns
    lines = response.split('\n')
    code_lines = []
    in_code = False
    for line in lines:
        stripped = line.strip()
        # Detect start of code
        if not in_code and stripped.startswith(('import ', 'from ', 'def ', 'class ', '# ', 'wb ', 'ws ', 'pd.', 'openpyxl')):
            in_code = True
        if in_code:
            # Stop at obvious non-code lines (numbered lists, prose)
            if stripped and stripped[0].isdigit() and '. ' in stripped[:5] and not stripped.startswith(('0', '1_', '2_', '3_')):
                # Numbered list item like "1. First step" - might be end of code
                # But "1_task_id" is a filename, keep it
                if not any(kw in stripped for kw in ['import', 'def', 'class', '=', '(', ')', '[', ']']):
                    break
            code_lines.append(line)
    
    if code_lines:
        return '\n'.join(code_lines).strip()
    
    # Last resort: return everything
    return response.strip()


def execute_code(code: str, work_dir: str, timeout: int = 120) -> tuple:
    """Execute Python code in a subprocess. Returns (success, output/error)."""
    code_path = os.path.join(work_dir, "solution.py")
    with open(code_path, "w") as f:
        f.write(code)

    try:
        result = subprocess.run(
            [sys.executable, code_path],
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=work_dir,
        )
        if result.returncode == 0:
            return True, result.stdout
        else:
            return False, result.stderr
    except subprocess.TimeoutExpired:
        return False, "Timeout"
    except Exception as e:
        return False, str(e)


# ============================================================
# Evaluation (from SpreadsheetBench)
# ============================================================

import datetime


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        excel_start = datetime.datetime(1899, 12, 30)
        delta = v - excel_start
        v = round(delta.days + delta.seconds / 86400.0, 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    return v1 == v2


def col_name2num(name):
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def col_num2name(n):
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def generate_cell_names(range_str):
    if ":" not in range_str:
        return [range_str]
    start, end = range_str.split(":")
    sc, sr = "", ""
    for c in start:
        if c.isdigit():
            sr += c
        else:
            sc += c
    ec, er = "", ""
    for c in end:
        if c.isdigit():
            er += c
        else:
            ec += c
    cols = [col_num2name(i) for i in range(col_name2num(sc), col_name2num(ec) + 1)]
    return [f"{col}{row}" for col in cols for row in range(int(sr), int(er) + 1)]


def evaluate_task(golden_path: str, output_path: str, answer_position: str) -> bool:
    """Evaluate agent output against golden answer. Returns True if pass."""
    if not os.path.exists(output_path):
        return False

    try:
        wb_gt = openpyxl.load_workbook(golden_path, data_only=True)
        wb_out = openpyxl.load_workbook(output_path, data_only=True)
    except Exception:
        return False

    # Normalize malformed formats like 'Sheet1!'A1 -> 'Sheet1'!A1
    answer_position = re.sub(r"'([^'!]+)!'", r"'\1'!", answer_position)
    answer_position = answer_position.rstrip("'").rstrip('"')
    
    parts = re.split(r",(?='|[A-Z])", answer_position)

    for part in parts:
        part = part.strip()
        if "!" in part:
            sheet_name, cell_range = part.rsplit("!", 1)
            sheet_name = sheet_name.strip("'").strip('"')
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = part

        cell_range = cell_range.strip("'").strip('"')

        if sheet_name not in wb_out.sheetnames:
            return False

        ws_gt = wb_gt[sheet_name]
        ws_out = wb_out[sheet_name]

        for cell_name in generate_cell_names(cell_range):
            if not compare_cell_value(ws_gt[cell_name].value, ws_out[cell_name].value):
                return False

    return True


# ============================================================
# Main experiment loop
# ============================================================

def get_task_family(task_id: str, pilot_split: dict) -> str:
    """Look up the family name for a task from the split file."""
    for key, fam in pilot_split.items():
        if not isinstance(fam, dict) or "tasks" not in fam:
            continue
        tasks = [str(t) for t in fam["tasks"]]
        if str(task_id) in tasks:
            return fam["name"]
    return "unknown"


def run_single(task_id, condition, artifact_type, family, model, api_key, base_url):
    """Run a single experiment. Returns result dict."""
    # Determine paths based on condition
    if condition == "original":
        spread_dir = os.path.join(VERIFIED_DIR, "spreadsheet", str(task_id))
        input_file = f"1_{task_id}_init.xlsx"
        golden_file = f"1_{task_id}_golden.xlsx"
        with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
            dataset = json.load(f)
        meta = next(d for d in dataset if str(d["id"]) == str(task_id))
    else:
        # Use correct perturbed dir based on task origin
        if str(task_id) in EXPANSION_TASKS:
            pdir = EXPANSION_PERTURBED_DIR
        else:
            pdir = PERTURBED_DIR
        spread_dir = os.path.join(pdir, str(task_id))
        input_file = f"1_{task_id}_init.xlsx"
        golden_file = f"1_{task_id}_golden.xlsx"
        with open(os.path.join(spread_dir, "task_meta.json")) as f:
            meta = json.load(f)

    input_path = os.path.join(spread_dir, input_file)
    golden_path = os.path.join(spread_dir, golden_file)

    if not os.path.exists(input_path):
        return {"task_id": task_id, "condition": condition, "artifact_type": artifact_type,
                "family": family, "pass": False, "error": "input file not found"}

    # Create temp work directory
    with tempfile.TemporaryDirectory() as work_dir:
        # Copy input file to work dir
        shutil.copy2(input_path, os.path.join(work_dir, input_file))
        output_file = f"1_{task_id}_output.xlsx"
        output_path = os.path.join(work_dir, output_file)

        # Build prompt
        prompt = build_prompt(
            task_meta=meta,
            input_path=os.path.join(work_dir, input_file),
            output_path=output_path,
            artifact_type=artifact_type,
            family=family,
            task_id=task_id,
        )

        # Call LLM
        try:
            response = call_llm(prompt, model, api_key, base_url)
        except Exception as e:
            return {"task_id": task_id, "condition": condition, "artifact_type": artifact_type,
                    "family": family, "pass": False, "error": f"LLM error: {e}"}

        # Extract and execute code
        code = extract_code(response)
        success, exec_output = execute_code(code, work_dir)

        # Evaluate
        passed = False
        if success and os.path.exists(output_path):
            passed = evaluate_task(golden_path, output_path, meta["answer_position"])

        return {
            "task_id": task_id,
            "condition": condition,
            "artifact_type": artifact_type,
            "family": family,
            "pass": passed,
            "code_executed": success,
            "model": model,
            "code": code[:500],
            "exec_output": exec_output[:200] if exec_output else "",
            "error": exec_output[:200] if not success else None,
            "output_file_exists": os.path.exists(output_path),
        }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", default="us.anthropic.claude-sonnet-4-5-20250929-v1:0")
    parser.add_argument("--api_key", default=None, help="OpenAI API key (not needed for Bedrock)")
    parser.add_argument("--base_url", default=None)
    parser.add_argument("--task_split", default="scratch/expanded-task-split.json",
                        help="Path to task split JSON")
    parser.add_argument("--task_id", default=None, help="Run single task only")
    parser.add_argument("--dry_run", action="store_true", help="Print prompts without calling LLM")
    parser.add_argument("--workers", type=int, default=8, help="Number of parallel workers")
    parser.add_argument("--repeats", type=int, default=1, help="Number of repeats per run")
    args = parser.parse_args()

    with open(args.task_split) as f:
        pilot_split = json.load(f)

    all_tasks = []
    for fam in pilot_split.values():
        if isinstance(fam, dict) and "tasks" in fam:
            all_tasks.extend([str(t) for t in fam["tasks"]])
    all_tasks = [t for t in all_tasks if t not in EXCLUDED_TASKS]

    if args.task_id:
        all_tasks = [args.task_id]

    os.makedirs(RESULTS_DIR, exist_ok=True)

    # Build job list
    jobs = []
    for task_id in all_tasks:
        family = get_task_family(task_id, pilot_split)
        for condition in CONDITIONS:
            for artifact_type in ARTIFACT_TYPES:
                for repeat in range(args.repeats):
                    jobs.append((str(task_id), condition, artifact_type, family, repeat))

    # Resume support: load existing results and skip completed jobs
    results_path = os.path.join(RESULTS_DIR, f"results_expanded_{args.model}.json")
    existing_results = []
    completed_keys = set()
    if os.path.exists(results_path):
        with open(results_path) as f:
            existing_results = json.load(f)
        for r in existing_results:
            completed_keys.add((str(r["task_id"]), r["condition"], r["artifact_type"], r.get("repeat", 0)))
        print(f"Resuming: {len(completed_keys)} runs already completed, {len(jobs) - len(completed_keys)} remaining")

    remaining_jobs = [j for j in jobs if (j[0], j[1], j[2], j[4]) not in completed_keys]

    if args.dry_run:
        for i, (task_id, condition, artifact_type, family, repeat) in enumerate(remaining_jobs):
            print(f"[{i+1}/{len(remaining_jobs)}] {task_id} | {condition} | {artifact_type} | r{repeat} (dry run)")
        return

    # Parallel execution
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    results = list(existing_results)  # start with existing
    lock = threading.Lock()
    done_count = [len(completed_keys)]
    total = len(jobs)

    def run_job(job):
        task_id, condition, artifact_type, family, repeat = job
        result = run_single(task_id, condition, artifact_type, family,
                            args.model, args.api_key, args.base_url)
        with lock:
            done_count[0] += 1
            status = "✅" if result["pass"] else "❌"
            extra = ""
            if not result.get("code_executed"):
                extra = f" [exec: {str(result.get('error', ''))[:60]}]"
            elif not result.get("output_file_exists"):
                extra = " [no output]"
            elif not result["pass"]:
                extra = " [wrong]"
            print(f"[{done_count[0]}/{total}] {status} {task_id} | {condition} | {artifact_type} | r{repeat}{extra}")
        result['repeat'] = repeat
        return result

    print(f"Running {len(remaining_jobs)} jobs with {args.workers} workers...")

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(run_job, job): job for job in remaining_jobs}
        for future in as_completed(futures):
            try:
                result = future.result()
                with lock:
                    results.append(result)
                    # Save incrementally after each result
                    with open(results_path, "w") as f:
                        json.dump(results, f, indent=2)
            except Exception as e:
                job = futures[future]
                print(f"ERROR on {job}: {e}")
                with lock:
                    results.append({
                        "task_id": job[0], "condition": job[1],
                        "artifact_type": job[2], "family": job[3],
                        "pass": False, "error": str(e),
                    })

    # Final save
    with open(results_path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"\nResults saved to {results_path}")

    # Print summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    for artifact_type in ARTIFACT_TYPES:
        for condition in CONDITIONS:
            subset = [r for r in results
                      if r["artifact_type"] == artifact_type and r["condition"] == condition]
            if subset:
                pass_rate = sum(r["pass"] for r in subset) / len(subset)
                print(f"  {artifact_type:20s} | {condition:10s} | {pass_rate:.1%} ({sum(r['pass'] for r in subset)}/{len(subset)})")


if __name__ == "__main__":
    main()
