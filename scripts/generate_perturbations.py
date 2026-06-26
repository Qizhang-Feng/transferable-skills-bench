#!/usr/bin/env python3
"""
Perturbation generator for pilot tasks.
Pipeline: load → rename workbook → rewrite instruction (rename) → layout shift → rewrite instruction (layout) → recompute answer_position → validate → save.
"""

import json
import os
import re
import copy
import shutil
from dataclasses import dataclass, field, asdict
from typing import Optional
import openpyxl


# ============================================================
# 1. Data structures
# ============================================================

@dataclass
class AnswerRef:
    sheet: Optional[str]
    start_col: str
    start_row: int
    end_col: Optional[str] = None
    end_row: Optional[int] = None

    def is_range(self):
        return self.end_col is not None


def parse_single_ref(s: str) -> AnswerRef:
    """Parse 'Sheet1'!A1:B10 or A1:B10 or A1 into AnswerRef.
    Also handles malformed formats like 'Sheet1!'A1:A50 (quote before !)."""
    s = s.strip().strip("'").strip('"')
    sheet = None
    cell_part = s

    # Handle sheet reference - multiple formats
    if '!' in s:
        sheet_part, cell_part = s.rsplit('!', 1)
        sheet = sheet_part.strip("'").strip('"')

    cell_part = cell_part.strip("'").strip('"')

    if ':' in cell_part:
        start, end = cell_part.split(':')
        sc, sr = _split_cell(start)
        ec, er = _split_cell(end)
        return AnswerRef(sheet=sheet, start_col=sc, start_row=sr, end_col=ec, end_row=er)
    else:
        sc, sr = _split_cell(cell_part)
        return AnswerRef(sheet=sheet, start_col=sc, start_row=sr)


def _split_cell(cell: str):
    """Split 'AB12' into ('AB', 12)."""
    col, row = '', ''
    for c in cell:
        if c.isalpha():
            col += c
        else:
            row += c
    return col, int(row) if row else 1


def parse_answer_position(s: str) -> list:
    """Parse comma-separated answer position string into list of AnswerRef."""
    # Normalize malformed formats like 'Sheet1!'A1 -> 'Sheet1'!A1
    s = re.sub(r"'([^'!]+)!'", r"'\1'!", s)
    # Remove trailing stray quotes
    s = s.rstrip("'").rstrip('"')
    
    refs = []
    # Split by comma, handling quoted sheet names
    # Simple approach: split on comma followed by a quote or letter
    parts = re.split(r",(?='|[A-Z])", s)
    for part in parts:
        part = part.strip()
        if part:
            refs.append(parse_single_ref(part))
    return refs


def apply_sheet_rename_to_refs(refs: list, sheet_map: dict) -> list:
    for r in refs:
        if r.sheet and r.sheet in sheet_map:
            r.sheet = sheet_map[r.sheet]
    return refs


def apply_row_shift_to_refs(refs: list, shift: int, sheets_to_shift: set = None) -> list:
    for r in refs:
        if sheets_to_shift:
            # If ref has no sheet name, it refers to the first/default sheet - shift it if any sheet was shifted
            if r.sheet is not None and r.sheet not in sheets_to_shift:
                continue
        r.start_row += shift
        if r.end_row is not None:
            r.end_row += shift
    return refs


def serialize_answer_position(refs: list) -> str:
    parts = []
    for r in refs:
        cell = f"{r.start_col}{r.start_row}"
        if r.is_range():
            cell += f":{r.end_col}{r.end_row}"
        if r.sheet:
            cell = f"'{r.sheet}'!{cell}"
        parts.append(cell)
    return ','.join(parts)


# ============================================================
# 2. Workbook rename
# ============================================================

def rename_workbook(wb, spec, schema):
    """Apply rename operations to a workbook. Returns log dict."""
    log = {'sheets_renamed': {}, 'headers_renamed': {}, 'data_values_renamed': [], 'formulas_updated': 0}

    sheet_map = spec['rename'].get('sheet_rename_map', {})
    col_map = spec['rename'].get('column_rename_map', {})
    data_renames = spec['rename'].get('data_value_renames', [])
    update_formulas = spec['rename'].get('formula_sheet_refs_to_update', False)

    # A. Sheet rename
    for old_name, new_name in sheet_map.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name
            log['sheets_renamed'][old_name] = new_name

    # B. Header rename - only on header row
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Find the schema entry (may be under old or new name)
        schema_entry = None
        for orig_sn, si in schema.get('sheets', {}).items():
            actual_name = sheet_map.get(orig_sn, orig_sn)
            if actual_name == sn:
                schema_entry = si
                break
        if not schema_entry:
            continue

        hdr_row = schema_entry.get('header_row')
        if hdr_row is None:
            continue

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=hdr_row, column=col_idx)
            if cell.value and str(cell.value) in col_map:
                old_val = str(cell.value)
                cell.value = col_map[old_val]
                log['headers_renamed'][f"{sn}!{cell.coordinate}"] = f"{old_val} -> {col_map[old_val]}"

    # C. Data value rename
    for dr in data_renames:
        find_text = dr['find']
        replace_text = dr['replace']
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and find_text in cell.value:
                        cell.value = cell.value.replace(find_text, replace_text)
                        log['data_values_renamed'].append(f"{sn}!{cell.coordinate}: {find_text} -> {replace_text}")

    # D. Formula sheet-ref update
    if update_formulas:
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        original = cell.value
                        for old_sn, new_sn in sheet_map.items():
                            # Replace sheet references in formulas
                            cell.value = cell.value.replace(f"'{old_sn}'!", f"'{new_sn}'!")
                            cell.value = cell.value.replace(f"{old_sn}!", f"{new_sn}!")
                        if cell.value != original:
                            log['formulas_updated'] += 1

    return log


# ============================================================
# 3. Instruction rewrite
# ============================================================

def rewrite_instruction_rename(text: str, spec: dict) -> str:
    """Rewrite instruction text for rename perturbation."""
    rules = spec['rename'].get('instruction_rewrite_rules', [])
    use_exact = spec['rename'].get('use_exact_match', False)

    if rules:
        for rule in rules:
            find = rule['find']
            replace = rule['replace']
            case_insensitive = rule.get('case_insensitive', False)
            if case_insensitive:
                text = re.sub(re.escape(find), replace, text, flags=re.IGNORECASE)
            else:
                text = text.replace(find, replace)
    else:
        # Fallback: apply column_rename_map and sheet_rename_map directly
        if not use_exact:
            for old, new in spec['rename'].get('column_rename_map', {}).items():
                text = text.replace(old, new)
            for old, new in spec['rename'].get('sheet_rename_map', {}).items():
                # Only replace quoted or clearly sheet-reference contexts
                text = text.replace(f"'{old}'", f"'{new}'")
                text = text.replace(f'"{old}"', f'"{new}"')

    return text


def rewrite_instruction_layout(text: str, spec: dict) -> str:
    """Rewrite instruction text for layout perturbation (row number shifts)."""
    shifts = spec['layout'].get('instruction_row_shifts', {})
    fixed_refs = spec['layout'].get('fixed_refs_to_shift', [])

    for old, new in shifts.items():
        text = text.replace(old, new)

    for ref_shift in fixed_refs:
        if ' -> ' in ref_shift:
            old_ref, new_ref = ref_shift.split(' -> ')
            text = text.replace(old_ref.strip(), new_ref.strip())

    return text


# ============================================================
# 4. Layout shift
# ============================================================

def apply_layout_shift(wb, layout_spec, schema, sheet_map):
    """Insert title row at row 1 for applicable sheets."""
    if layout_spec.get('skip_layout', False):
        return []

    title_text = layout_spec.get('title_text', 'Report')
    selective = layout_spec.get('selective_sheets', None)
    shifted_sheets = []

    for sn in wb.sheetnames:
        if selective and sn not in selective:
            continue

        ws = wb[sn]
        ws.insert_rows(1, amount=1)
        ws.cell(row=1, column=1, value=title_text)
        shifted_sheets.append(sn)

    return shifted_sheets


# ============================================================
# 5. Validation
# ============================================================

def validate_task(wb_init, wb_golden, perturbed_instruction, perturbed_ap, spec, schema):
    """Run validation checks. Returns dict of check results."""
    results = {}

    sheet_map = spec['rename'].get('sheet_rename_map', {})

    # Check 1: All sheets in answer_position exist in workbook
    refs = parse_answer_position(perturbed_ap)
    for r in refs:
        if r.sheet:
            results[f'sheet_exists_{r.sheet}'] = r.sheet in wb_init.sheetnames

    # Check 2: Sheet names are unique
    results['sheet_names_unique'] = len(wb_init.sheetnames) == len(set(wb_init.sheetnames))

    # Check 3: Instruction doesn't contain old sheet names (word-boundary check)
    for old_sn in sheet_map.keys():
        if len(old_sn) > 2:
            # Use word boundary to avoid false positives (e.g. Data inside SourceData)
            if re.search(r'(?<![A-Za-z])' + re.escape(old_sn) + r'(?![A-Za-z])', perturbed_instruction):
                results[f'instruction_still_has_{old_sn}'] = True
            else:
                results[f'instruction_still_has_{old_sn}'] = False

    # Check 4: Instruction doesn't contain old column names (word-boundary check)
    for old_col in spec['rename'].get('column_rename_map', {}).keys():
        if len(old_col) > 2:
            # Use word boundary to avoid false positives (e.g. SALE inside SALES_QTY)
            if re.search(r'\b' + re.escape(old_col) + r'\b', perturbed_instruction):
                results[f'instruction_still_has_col_{old_col}'] = True

    return results


# ============================================================
# 6. Main pipeline
# ============================================================

def process_task(task_id, spec, schema, task_meta, src_dir, out_dir):
    """Process one task: rename + layout + validate + save."""
    print(f"\n{'='*60}")
    print(f"Processing {task_id} [{spec.get('family', '?')}]")
    print(f"{'='*60}")

    init_path = os.path.join(src_dir, str(task_id), f"1_{task_id}_init.xlsx")
    golden_path = os.path.join(src_dir, str(task_id), f"1_{task_id}_golden.xlsx")
    prompt_path = os.path.join(src_dir, str(task_id), "prompt.txt")

    # Load
    wb_init = openpyxl.load_workbook(init_path)
    wb_golden = openpyxl.load_workbook(golden_path)
    # Also load golden with data_only=True to get cached formula values
    wb_golden_data = openpyxl.load_workbook(golden_path, data_only=True)
    instruction = task_meta['instruction']
    original_ap = task_meta['answer_position']

    # Step 0: Bake formulas in golden BEFORE any modifications
    # Replace formula cells with their cached values so they survive save/load
    from openpyxl.worksheet.formula import ArrayFormula
    formulas_baked = 0
    formulas_unbaked = 0
    for sn in wb_golden.sheetnames:
        ws = wb_golden[sn]
        ws_data = wb_golden_data[sn]
        for row in ws.iter_rows():
            for cell in row:
                is_formula = False
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    is_formula = True
                elif isinstance(cell.value, ArrayFormula):
                    is_formula = True
                
                if is_formula:
                    cached = ws_data[cell.coordinate].value
                    if cached is not None:
                        cell.value = cached
                        formulas_baked += 1
                    else:
                        formulas_unbaked += 1
    if formulas_baked > 0:
        print(f"  Baked {formulas_baked} formula cells in golden to static values")
    if formulas_unbaked > 0:
        print(f"  ⚠️ {formulas_unbaked} formula cells could NOT be baked (no cached value)")

    # Parse original answer_position
    refs = parse_answer_position(original_ap)
    sheet_map = spec['rename'].get('sheet_rename_map', {})

    # Step 1: Rename workbooks
    init_rename_log = rename_workbook(wb_init, spec, schema)
    golden_rename_log = rename_workbook(wb_golden, spec, schema)
    print(f"  Rename: {len(init_rename_log['sheets_renamed'])} sheets, "
          f"{len(init_rename_log['headers_renamed'])} headers, "
          f"{len(init_rename_log['data_values_renamed'])} data values, "
          f"{init_rename_log['formulas_updated']} formulas")

    # Step 2: Rewrite instruction for rename
    perturbed_instruction = rewrite_instruction_rename(instruction, spec)

    # Step 3: Apply sheet rename to answer refs
    refs = apply_sheet_rename_to_refs(refs, sheet_map)

    # Step 4: Layout shift
    layout_spec = spec.get('layout', {})
    shifted_sheets = apply_layout_shift(wb_init, layout_spec, schema, sheet_map)
    apply_layout_shift(wb_golden, layout_spec, schema, sheet_map)
    print(f"  Layout: shifted {len(shifted_sheets)} sheets: {shifted_sheets}")

    # Step 5: Rewrite instruction for layout
    if shifted_sheets:
        perturbed_instruction = rewrite_instruction_layout(perturbed_instruction, spec)

    # Step 6: Apply row shift to answer refs
    if shifted_sheets:
        refs = apply_row_shift_to_refs(refs, shift=1, sheets_to_shift=set(shifted_sheets))

    # Serialize
    # Check for manual override (e.g., when answer sheet doesn't exist in init)
    ap_override = spec.get('layout', {}).get('answer_position_override')
    if ap_override:
        perturbed_ap = ap_override
        print(f"  Answer position (override): {original_ap} -> {perturbed_ap}")
    else:
        perturbed_ap = serialize_answer_position(refs)
        print(f"  Answer position: {original_ap} -> {perturbed_ap}")

    # Step 7: Validate
    validation = validate_task(wb_init, wb_golden, perturbed_instruction, perturbed_ap, spec, schema)
    issues = {k: v for k, v in validation.items() if v is True and 'still_has' in k}
    if issues:
        print(f"  ⚠️ Validation issues: {issues}")
    else:
        print(f"  ✅ Validation passed")

    # Step 9: Save
    task_out = os.path.join(out_dir, str(task_id))
    os.makedirs(task_out, exist_ok=True)

    wb_init.save(os.path.join(task_out, f"1_{task_id}_init.xlsx"))
    wb_golden.save(os.path.join(task_out, f"1_{task_id}_golden.xlsx"))

    # Save manifest
    manifest = {
        'task_id': task_id,
        'family': spec.get('family', ''),
        'status': 'ok' if not issues else 'ok_with_warnings',
        'applied': {
            'rename': bool(sheet_map or spec['rename'].get('column_rename_map')),
            'layout': bool(shifted_sheets),
        },
        'instruction_original': instruction[:200] + '...',
        'instruction_perturbed': perturbed_instruction[:200] + '...',
        'answer_position_original': original_ap,
        'answer_position_perturbed': perturbed_ap,
        'sheet_renames': sheet_map,
        'header_renames': spec['rename'].get('column_rename_map', {}),
        'data_value_renames': spec['rename'].get('data_value_renames', []),
        'layout_shifted_sheets': shifted_sheets,
        'validation': validation,
        'rename_log': init_rename_log,
    }

    with open(os.path.join(task_out, 'manifest.json'), 'w') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    # Also save full perturbed instruction
    with open(os.path.join(task_out, 'instruction_perturbed.txt'), 'w') as f:
        f.write(perturbed_instruction)

    # Save perturbed metadata
    perturbed_meta = {
        'id': task_id,
        'instruction': perturbed_instruction,
        'instruction_type': task_meta['instruction_type'],
        'answer_position': perturbed_ap,
    }
    with open(os.path.join(task_out, 'task_meta.json'), 'w') as f:
        json.dump(perturbed_meta, f, indent=2, ensure_ascii=False)

    return manifest


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--spec_path', default='scratch/pilot-perturbation-specs-v3.json')
    parser.add_argument('--schema_path', default='scratch/pilot-task-schemas-v2.json')
    parser.add_argument('--out_dir', default='scratch/pilot_perturbed')
    args = parser.parse_args()
    
    SPEC_PATH = args.spec_path
    SCHEMA_PATH = args.schema_path
    DATASET_PATH = 'vendor/SpreadsheetBench/data/spreadsheetbench_verified_400/dataset.json'
    SRC_DIR = 'vendor/SpreadsheetBench/data/spreadsheetbench_verified_400/spreadsheet'
    OUT_DIR = args.out_dir

    with open(SPEC_PATH) as f:
        specs = json.load(f)
    with open(SCHEMA_PATH) as f:
        schemas = json.load(f)
    with open(DATASET_PATH) as f:
        dataset = json.load(f)

    task_meta_lookup = {str(d['id']): d for d in dataset}

    # Filter usable tasks
    usable = {tid: s for tid, s in specs.items() if s.get('review_status') in ('ok', 'fixed')}
    print(f"Processing {len(usable)} usable tasks")

    os.makedirs(OUT_DIR, exist_ok=True)

    all_manifests = []
    for tid, spec in usable.items():
        schema = schemas.get(tid, {})
        meta = task_meta_lookup.get(tid, {})
        if not meta:
            print(f"WARNING: no metadata for {tid}, skipping")
            continue

        manifest = process_task(tid, spec, schema, meta, SRC_DIR, OUT_DIR)
        all_manifests.append(manifest)

    # Save combined manifest
    with open(os.path.join(OUT_DIR, 'all_manifests.json'), 'w') as f:
        json.dump(all_manifests, f, indent=2, ensure_ascii=False)

    # Summary
    print(f"\n{'='*60}")
    print(f"SUMMARY: {len(all_manifests)} tasks processed")
    ok = sum(1 for m in all_manifests if m['status'] == 'ok')
    warn = sum(1 for m in all_manifests if m['status'] == 'ok_with_warnings')
    print(f"  OK: {ok}, OK with warnings: {warn}")


if __name__ == '__main__':
    main()
