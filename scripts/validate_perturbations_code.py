#!/usr/bin/env python3
"""
Layer 1: Deterministic code validator for perturbation specs.
Checks all P0 hard rules that don't need LLM judgment.

Criteria covered:
- Rename namespace validity & uniqueness (C3)
- Answer range mapping correctness (C11)
- Answer range validity against schema (C12)
- Instruction completeness: old name leftovers (C5)
- Instruction/answer/schema consistency (C9)
- Rewrite collision with common English (C7)
- Rewrite boundary / partial-match (C8)
- Formula baking coverage (C15)
- Range shape preservation
- Header row assumption check
"""

import json
import os
import re
import openpyxl

VERIFIED_DIR = "vendor/SpreadsheetBench/data/spreadsheetbench_verified_400"

# Common English words that should NOT be replaced case-insensitively
COMMON_ENGLISH_TOKENS = {
    "name", "names", "named", "date", "dates", "dated",
    "may", "week", "weeks", "weekly", "sum", "total", "totals",
    "data", "type", "types", "rate", "rates", "rated",
    "key", "keys", "code", "codes", "id", "ids",
    "amount", "amounts", "value", "values", "result", "results",
    "entry", "entries", "record", "records", "score", "scores",
    "group", "groups", "grouped", "item", "items",
    "average", "max", "min", "count",
}

# Excel illegal sheet name characters
EXCEL_ILLEGAL_SHEET_CHARS = set("\\/*?:[]")
EXCEL_MAX_SHEET_NAME_LEN = 31


def check_rename_namespace(spec, schema):
    """C3: Rename namespace validity and uniqueness."""
    issues = []
    col_map = spec["rename"].get("column_rename_map", {})
    sheet_map = spec["rename"].get("sheet_rename_map", {})

    # Check sheet name legality
    for old, new in sheet_map.items():
        if len(new) > EXCEL_MAX_SHEET_NAME_LEN:
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Sheet name '{new}' exceeds Excel 31-char limit ({len(new)} chars)",
            })
        if any(c in new for c in EXCEL_ILLEGAL_SHEET_CHARS):
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Sheet name '{new}' contains illegal Excel characters",
            })
        # Check if new name looks like a cell reference
        if re.match(r'^[A-Z]{1,3}\d+$', new):
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Sheet name '{new}' looks like a cell reference",
            })

    # Check sheet name uniqueness after rename
    sheet_names = list(schema.get("sheet_names", []))
    renamed_sheets = []
    for sn in sheet_names:
        renamed_sheets.append(sheet_map.get(sn, sn))
    if len(renamed_sheets) != len(set(renamed_sheets)):
        dupes = [s for s in renamed_sheets if renamed_sheets.count(s) > 1]
        issues.append({
            "severity": "P0", "criterion": "rename_namespace",
            "description": f"Sheet name collision after rename: {set(dupes)}",
        })

    # Check header uniqueness after rename within each sheet
    for sh in schema.get("sheets", []):
        headers = [h["value"] for h in sh.get("headers", [])]
        renamed_headers = [col_map.get(h, h) for h in headers]
        if len(renamed_headers) != len(set(renamed_headers)):
            dupes = [h for h in renamed_headers if renamed_headers.count(h) > 1]
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Header collision in sheet '{sh['name']}' after rename: {set(dupes)}",
            })

    # Check that rename doesn't map two different old names to the same new name
    new_col_names = list(col_map.values())
    if len(new_col_names) != len(set(new_col_names)):
        dupes = [n for n in new_col_names if new_col_names.count(n) > 1]
        issues.append({
            "severity": "P0", "criterion": "rename_namespace",
            "description": f"Column rename maps multiple old names to same new name: {set(dupes)}",
        })

    return issues


def check_answer_range_mapping(spec, meta, perturbed_meta):
    """C11: Answer range mapping correctness."""
    issues = []
    orig_ap = meta.get("answer_position", "")
    pert_ap = perturbed_meta.get("answer_position", "")
    sheet_map = spec["rename"].get("sheet_rename_map", {})

    if not orig_ap or not pert_ap:
        issues.append({
            "severity": "P0", "criterion": "answer_range_mapping",
            "description": "Missing answer position (original or perturbed)",
        })
        return issues

    # Check sheet name update in answer position
    for old_sheet, new_sheet in sheet_map.items():
        if old_sheet in orig_ap and new_sheet not in pert_ap:
            issues.append({
                "severity": "P0", "criterion": "answer_range_mapping",
                "description": f"Sheet '{old_sheet}' in original AP but '{new_sheet}' not in perturbed AP",
            })
        if old_sheet in pert_ap:
            issues.append({
                "severity": "P0", "criterion": "answer_range_mapping",
                "description": f"Old sheet name '{old_sheet}' still in perturbed AP: {pert_ap}",
            })

    # Check row shift: extract row numbers and verify +1
    orig_rows = re.findall(r'(\d+)', orig_ap)
    pert_rows = re.findall(r'(\d+)', pert_ap)

    if len(orig_rows) == len(pert_rows) and len(orig_rows) > 0:
        all_shifted = all(
            int(p) == int(o) + 1
            for o, p in zip(orig_rows, pert_rows)
        )
        no_shift = all(
            int(p) == int(o)
            for o, p in zip(orig_rows, pert_rows)
        )
        if not all_shifted and not no_shift:
            issues.append({
                "severity": "P1", "criterion": "answer_range_mapping",
                "description": f"Row shift inconsistent. Orig rows: {orig_rows}, Pert rows: {pert_rows}. Expected all +1 or all same.",
            })

    # Check range shape preservation
    def parse_range_shape(ap):
        """Extract (rows, cols) from a range like B3:C68."""
        ranges = re.findall(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', ap)
        shapes = []
        for c1, r1, c2, r2 in ranges:
            rows = int(r2) - int(r1) + 1
            cols = ord(c2) - ord(c1) + 1
            shapes.append((rows, cols))
        return shapes

    orig_shapes = parse_range_shape(orig_ap)
    pert_shapes = parse_range_shape(pert_ap)
    if orig_shapes and pert_shapes and orig_shapes != pert_shapes:
        issues.append({
            "severity": "P0", "criterion": "range_shape_preservation",
            "description": f"Range shape changed. Orig: {orig_shapes}, Pert: {pert_shapes}",
        })

    return issues


def check_answer_range_validity(perturbed_meta, schema):
    """C12: Answer range validity against schema."""
    issues = []
    pert_ap = perturbed_meta.get("answer_position", "")

    # Extract max row from ranges
    row_refs = re.findall(r':?[A-Z]+(\d+)', pert_ap)
    if not row_refs:
        return issues

    max_row_in_ap = max(int(r) for r in row_refs)

    # Check against schema (after layout shift, schema rows + 1)
    for sh in schema.get("sheets", []):
        schema_rows = sh.get("rows", 0) + 1  # +1 for layout shift
        # Only check if the answer references this sheet
        sheet_name = sh.get("name", "")
        if sheet_name in pert_ap or not any("!" in pert_ap for _ in [1]):
            if max_row_in_ap > schema_rows + 10:  # Allow some margin for output rows
                issues.append({
                    "severity": "P1", "criterion": "answer_range_validity",
                    "description": f"Answer range max row ({max_row_in_ap}) >> schema rows ({schema_rows}) for sheet '{sheet_name}'. May be OK if task generates output rows.",
                })

    return issues


def check_instruction_old_name_leftovers(spec, perturbed_instruction):
    """C5: Check if old names still appear in perturbed instruction."""
    issues = []
    col_map = spec["rename"].get("column_rename_map", {})
    sheet_map = spec["rename"].get("sheet_rename_map", {})

    for old_name in list(col_map.keys()) + list(sheet_map.keys()):
        if len(old_name) <= 2:
            continue  # Skip very short names (too many false positives)
        # Check exact match (case-sensitive)
        if old_name in perturbed_instruction:
            # Check it's not inside the new name
            new_name = col_map.get(old_name, sheet_map.get(old_name, ""))
            if old_name not in new_name:
                issues.append({
                    "severity": "P0", "criterion": "instruction_completeness",
                    "description": f"Old name '{old_name}' still appears in perturbed instruction",
                })

    return issues


def check_instruction_answer_schema_consistency(spec, perturbed_instruction, perturbed_meta):
    """C9: Check consistency between instruction, answer position, and schema."""
    issues = []
    sheet_map = spec["rename"].get("sheet_rename_map", {})
    pert_ap = perturbed_meta.get("answer_position", "")

    # Extract sheet names from answer position
    ap_sheets = re.findall(r"'([^']+)'!", pert_ap)
    ap_sheets += re.findall(r"(\w+)!", pert_ap)
    ap_sheets = set(ap_sheets)

    # Check that answer position sheet names appear in instruction
    for sheet in ap_sheets:
        if len(sheet) > 3 and sheet not in perturbed_instruction:
            # Not necessarily an error - some tasks don't mention the answer sheet
            issues.append({
                "severity": "P1", "criterion": "instruction_answer_consistency",
                "description": f"Sheet '{sheet}' in answer position but not mentioned in instruction",
            })

    # Check old sheet names don't appear in perturbed instruction
    for old_sheet in sheet_map.keys():
        if len(old_sheet) > 2 and old_sheet in perturbed_instruction:
            new_sheet = sheet_map[old_sheet]
            if old_sheet not in new_sheet:
                issues.append({
                    "severity": "P0", "criterion": "instruction_answer_consistency",
                    "description": f"Old sheet name '{old_sheet}' still in perturbed instruction (should be '{new_sheet}')",
                })

    return issues


def check_rewrite_collision(spec, original_instruction):
    """C7+C8: Check rewrite rules for common English collision and partial matches."""
    issues = []
    rules = spec["rename"].get("instruction_rewrite_rules", [])

    for rule in rules:
        find_text = rule["find"]
        case_insensitive = rule.get("case_insensitive", False)

        # Check collision with common English
        if case_insensitive and find_text.lower() in COMMON_ENGLISH_TOKENS:
            issues.append({
                "severity": "P0", "criterion": "rewrite_collision",
                "description": f"Case-insensitive rule '{find_text}' collides with common English word '{find_text.lower()}'",
            })

        # Check partial match risk
        if case_insensitive and len(find_text) <= 4:
            issues.append({
                "severity": "P1", "criterion": "rewrite_boundary",
                "description": f"Short case-insensitive rule '{find_text}' ({len(find_text)} chars) has high partial-match risk",
            })

        # Check if find_text appears as substring in longer words in instruction
        if case_insensitive:
            pattern = re.compile(r'\w' + re.escape(find_text) + r'|\b' + re.escape(find_text) + r'\w', re.IGNORECASE)
            matches = pattern.findall(original_instruction)
            if matches:
                issues.append({
                    "severity": "P1", "criterion": "rewrite_boundary",
                    "description": f"Rule '{find_text}' may partial-match in: {matches[:3]}",
                })

    return issues


def check_formula_baking(task_id, perturbed_dir):
    """C15: Check that golden file has no remaining formulas in target cells."""
    issues = []
    golden_path = os.path.join(perturbed_dir, str(task_id), f"1_{task_id}_golden.xlsx")

    if not os.path.exists(golden_path):
        issues.append({
            "severity": "P0", "criterion": "formula_baking",
            "description": "Golden file not found",
        })
        return issues

    try:
        wb = openpyxl.load_workbook(golden_path)
        formula_count = 0
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb.close()

        if formula_count > 0:
            issues.append({
                "severity": "P0", "criterion": "formula_baking",
                "description": f"Golden file still has {formula_count} formula cells (should be 0 after baking)",
            })
    except Exception as e:
        issues.append({
            "severity": "P0", "criterion": "formula_baking",
            "description": f"Error reading golden file: {str(e)[:100]}",
        })

    return issues


def check_header_row_assumption(original_instruction):
    """Check if instruction references 'row 1', 'first row', etc."""
    issues = []
    patterns = [
        r'\brow\s*1\b', r'\bfirst\s+row\b', r'\btop\s+row\b',
        r'\bheader\s+row\b', r'\brow\s+1\b', r'\bA1\b', r'\bB1\b',
        r'\bstarting\s+at\s+row\s*1\b',
    ]
    for pat in patterns:
        if re.search(pat, original_instruction, re.IGNORECASE):
            issues.append({
                "severity": "P1", "criterion": "layout_shift_neutrality",
                "description": f"Instruction references '{re.search(pat, original_instruction, re.IGNORECASE).group()}' — layout shift may not be neutral",
            })

    return issues


def validate_task(task_id, spec, meta, schema, perturbed_dir):
    """Run all code-based validations for one task."""
    all_issues = []

    # Load perturbed metadata
    pert_meta_path = os.path.join(perturbed_dir, str(task_id), "task_meta.json")
    pert_inst_path = os.path.join(perturbed_dir, str(task_id), "instruction_perturbed.txt")

    if not os.path.exists(pert_meta_path):
        return [{"severity": "P0", "criterion": "file_missing",
                 "description": f"Perturbed task_meta.json not found for {task_id}"}]

    with open(pert_meta_path) as f:
        pert_meta = json.load(f)
    with open(pert_inst_path) as f:
        pert_inst = f.read()

    # Run all checks
    all_issues.extend(check_rename_namespace(spec, schema))
    all_issues.extend(check_answer_range_mapping(spec, meta, pert_meta))
    all_issues.extend(check_answer_range_validity(pert_meta, schema))
    all_issues.extend(check_instruction_old_name_leftovers(spec, pert_inst))
    all_issues.extend(check_instruction_answer_schema_consistency(spec, pert_inst, pert_meta))
    all_issues.extend(check_rewrite_collision(spec, meta.get("instruction", "")))
    all_issues.extend(check_formula_baking(task_id, perturbed_dir))
    all_issues.extend(check_header_row_assumption(meta.get("instruction", "")))

    return all_issues


def main():
    with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
        dataset = json.load(f)
    meta_map = {str(d["id"]): d for d in dataset}

    with open("scratch/expansion-perturbation-specs.json") as f:
        specs = json.load(f)

    with open("scratch/expansion-task-schemas.json") as f:
        schemas_list = json.load(f)
    schema_map = {s["task_id"]: s for s in schemas_list}

    perturbed_dir = "scratch/expansion_perturbed"

    all_results = []
    for tid in sorted(specs.keys()):
        spec = specs[tid]
        meta = meta_map.get(tid, {})
        schema = schema_map.get(tid, {})

        issues = validate_task(tid, spec, meta, schema, perturbed_dir)

        p0 = [i for i in issues if i["severity"] == "P0"]
        p1 = [i for i in issues if i["severity"] == "P1"]

        if p0:
            verdict = "FAIL"
        elif p1:
            verdict = "WARN"
        else:
            verdict = "PASS"

        symbol = {"PASS": "V", "WARN": "W", "FAIL": "X"}.get(verdict, "?")
        print(f"  {symbol} {tid:12s} | {verdict:4s} | P0={len(p0)} P1={len(p1)}")
        for i in p0:
            print(f"      [P0] {i['criterion']}: {i['description'][:100]}")
        for i in p1:
            print(f"      [P1] {i['criterion']}: {i['description'][:100]}")

        all_results.append({
            "task_id": tid,
            "verdict": verdict,
            "issues": issues,
        })

    # Summary
    verdicts = [r["verdict"] for r in all_results]
    print(f"\nSUMMARY: PASS={verdicts.count('PASS')} WARN={verdicts.count('WARN')} FAIL={verdicts.count('FAIL')}")

    with open("scratch/expansion-code-validation-results.json", "w") as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)
    print("Results saved to scratch/expansion-code-validation-results.json")


if __name__ == "__main__":
    main()
