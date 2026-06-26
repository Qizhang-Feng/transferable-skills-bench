"""
Stage 1: Extract structured schema from a task's spreadsheet files.
"""

import re
import openpyxl


def detect_header_row(ws, max_scan=10):
    """Find the first row with multiple non-empty text cells."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        text_cells = 0
        for cell in ws[row_idx]:
            if cell.value is not None and isinstance(cell.value, str):
                text_cells += 1
        if text_cells >= 2:
            return row_idx
    # Fallback: first row with any content
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        for cell in ws[row_idx]:
            if cell.value is not None:
                return row_idx
    return None


def get_headers(ws, header_row=None):
    """Get header values from the detected header row."""
    if header_row is None:
        header_row = detect_header_row(ws)
    if header_row is None:
        return []
    headers = []
    for cell in ws[header_row]:
        if cell.value is not None:
            headers.append({
                "name": str(cell.value),
                "col_letter": cell.column_letter,
                "col_idx": cell.column,
            })
    return headers


def count_formulas(ws):
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count


def count_formulas_in_range(ws, range_str):
    """Count formula cells within a specific range."""
    # Simple implementation: check all cells
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count


def find_data_header_overlaps(ws, headers):
    """Find headers that also appear as data values in the sheet."""
    header_names = {h["name"] for h in headers}
    overlaps = []
    header_row = headers[0]["col_idx"] if headers else 1  # approximate

    for row in ws.iter_rows(min_row=2):  # skip header
        for cell in row:
            if isinstance(cell.value, str) and cell.value in header_names:
                overlaps.append({
                    "header": cell.value,
                    "cell": f"{cell.column_letter}{cell.row}",
                })
    return overlaps


def parse_instruction_refs(instruction, sheet_names):
    """Extract references from instruction text."""
    refs = {
        "sheet_names_mentioned": [],
        "column_names_mentioned": [],
        "cell_refs": [],
        "row_refs": [],
        "absolute_refs": [],
    }

    # Sheet names
    for sn in sheet_names:
        if len(sn) <= 2:
            # Short names: require quotes
            if f"'{sn}'" in instruction or f'"{sn}"' in instruction:
                refs["sheet_names_mentioned"].append(sn)
        else:
            if sn.lower() in instruction.lower():
                refs["sheet_names_mentioned"].append(sn)

    # Cell references: A1, B2, $C$5, etc.
    for m in re.finditer(r'\$?[A-Z]{1,3}\$?\d+', instruction):
        ref = m.group()
        if "$" in ref:
            refs["absolute_refs"].append(ref)
        else:
            refs["cell_refs"].append(ref)

    # Row references: "row 1", "Row 14", "first row"
    for m in re.finditer(r'[Rr]ow\s+\d+', instruction):
        refs["row_refs"].append(m.group())
    for m in re.finditer(r'first\s+row', instruction, re.IGNORECASE):
        refs["row_refs"].append(m.group())

    return refs


def build_reference_inventory(wb):
    """Scan workbook for global references that rename might affect."""
    inventory = {
        "formula_sheet_refs": set(),
        "defined_names": [],
        "data_validations": 0,
        "charts": 0,
        "pivot_tables": 0,
        "merged_cells_total": 0,
    }

    for sn in wb.sheetnames:
        ws = wb[sn]
        inventory["merged_cells_total"] += len(ws.merged_cells.ranges)

        # Scan formulas for sheet references
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    for ref_sheet in re.findall(r"'([^']+)'!", cell.value):
                        inventory["formula_sheet_refs"].add(ref_sheet)
                    for ref_sheet in re.findall(r"(\w+)!", cell.value):
                        if ref_sheet not in ("TRUE", "FALSE"):
                            inventory["formula_sheet_refs"].add(ref_sheet)

        # Data validations
        if ws.data_validations:
            inventory["data_validations"] += len(ws.data_validations.dataValidation)

    inventory["formula_sheet_refs"] = list(inventory["formula_sheet_refs"])

    # Defined names
    if wb.defined_names:
        try:
            # openpyxl >= 3.1: DefinedNameDict is iterable
            for dn in wb.defined_names.values():
                inventory["defined_names"].append(str(dn.name))
        except (AttributeError, TypeError):
            try:
                # Older openpyxl: has .definedName list
                for dn in wb.defined_names.definedName:
                    inventory["defined_names"].append(str(dn.name))
            except (AttributeError, TypeError):
                pass

    return inventory


def extract_schema(task_id, init_path, golden_path, instruction, answer_position):
    """Full schema extraction for one task."""
    wb = openpyxl.load_workbook(init_path)
    wb_golden = openpyxl.load_workbook(golden_path)
    wb_golden_data = openpyxl.load_workbook(golden_path, data_only=True)

    sheets = {}
    all_headers = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdr_row = detect_header_row(ws)
        headers = get_headers(ws, hdr_row)
        all_headers.extend(headers)
        overlaps = find_data_header_overlaps(ws, headers) if headers else []

        sheets[sn] = {
            "rows": ws.max_row,
            "cols": ws.max_column,
            "header_row": hdr_row,
            "headers": headers,
            "formula_count": count_formulas(ws),
            "merged_cells": len(ws.merged_cells.ranges),
            "data_header_overlaps": overlaps,
        }

    # Golden formula info
    formula_total = 0
    formula_in_answer = 0
    unbakeable = 0
    for sn in wb_golden.sheetnames:
        ws_f = wb_golden[sn]
        ws_d = wb_golden_data[sn]
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_total += 1
                    cached = ws_d[cell.coordinate].value
                    if cached is None:
                        unbakeable += 1

    ref_inventory = build_reference_inventory(wb)
    instruction_refs = parse_instruction_refs(instruction, wb.sheetnames)

    # Add column names mentioned in instruction
    header_names = [h["name"] for h in all_headers if len(h["name"]) > 2]
    for hn in header_names:
        pattern = re.compile(r'\b' + re.escape(hn) + r'\b', re.IGNORECASE)
        if pattern.search(instruction):
            instruction_refs["column_names_mentioned"].append(hn)
    instruction_refs["column_names_mentioned"] = list(set(
        instruction_refs["column_names_mentioned"]))

    wb.close()
    wb_golden.close()
    wb_golden_data.close()

    return {
        "task_id": task_id,
        "sheets": sheets,
        "sheet_names": list(sheets.keys()),
        "instruction_refs": instruction_refs,
        "golden_info": {
            "formula_cells_total": formula_total,
            "unbakeable_in_answer": unbakeable,
        },
        "reference_inventory": ref_inventory,
    }
