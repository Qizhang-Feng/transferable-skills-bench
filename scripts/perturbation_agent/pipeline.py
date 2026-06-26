#!/usr/bin/env python3
"""
Auto-Perturbation Agent Pipeline v3
Main orchestrator: 400 tasks parallel, 7 LLM guardrails, feedback loop.
"""

import json
import os
import re
import hashlib
import shutil
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter

from .schema_extract import extract_schema
from .feasibility import check_feasibility
from .coord_map import CoordMap
from .llm_client import call_llm, call_llm_json
from .prompts import (
    FAMILY_CLASSIFY_PROMPT, DESIGN_PROMPT, RENAME_SAFETY_PROMPT,
    LAYOUT_IMPACT_PROMPT, REWRITE_INSTRUCTION_PROMPT,
    INSTRUCTION_QUALITY_PROMPT, FINAL_SANITY_PROMPT,
)

# ============================================================
# Config
# ============================================================

VERIFIED_DIR = "vendor/SpreadsheetBench/data/spreadsheetbench_verified_400"
OUTPUT_DIR = "data/perturbed_v3"
MAX_RETRIES = 3
LLM_MODEL = "us.anthropic.claude-opus-4-6-v1"


# ============================================================
# Stage functions (stubs to be filled)
# ============================================================

def stage_0_classify_family(instruction, schema):
    """LLM #0: Classify task family + identify risk flags."""
    prompt = FAMILY_CLASSIFY_PROMPT.format(
        instruction=instruction[:1000],
        sheet_names=schema["sheet_names"],
        schema_summary=_format_schema_summary(schema),
    )
    result = call_llm_json(prompt, model=LLM_MODEL)
    return result.get("family", "unknown"), result.get("risk_flags", [])


def stage_3_design_perturbation(schema, instruction, answer_position,
                                 family, family_guide, safe_anchors,
                                 risk_flags, error_history):
    """LLM #1: Design best-fit perturbation spec."""
    error_section = _format_error_history(error_history)
    prompt = DESIGN_PROMPT.format(
        family_guide=family_guide,
        instruction=instruction[:1500],
        answer_position=answer_position,
        schema_json=json.dumps(schema["sheets"], indent=2, default=str)[:3000],
        safe_anchors=safe_anchors,
        risk_flags=risk_flags,
        error_section=error_section,
    )
    return call_llm_json(prompt, model=LLM_MODEL)


def stage_3_5_check_rename_safety(spec, instruction, schema):
    """LLM #1.5: Guardrail — check rename neutrality."""
    rename = spec.get("rename", {})
    if not rename.get("column_rename_map") and not rename.get("sheet_rename_map"):
        return {"approved": True, "issues": []}

    all_headers = {}
    for sn, info in schema["sheets"].items():
        all_headers[sn] = [h["name"] for h in info.get("headers", [])]

    prompt = RENAME_SAFETY_PROMPT.format(
        column_rename_map=json.dumps(rename.get("column_rename_map", {})),
        sheet_rename_map=json.dumps(rename.get("sheet_rename_map", {})),
        data_value_renames=json.dumps(rename.get("data_value_renames", [])),
        instruction=instruction[:500],
        all_headers=json.dumps(all_headers),
    )
    return call_llm_json(prompt, model=LLM_MODEL)


def stage_4_apply_perturbation(task, spec, schema, formula_policy):
    """Code: Apply perturbation to files. Returns (coord_map, perturbed_paths)."""
    coord_map = CoordMap()

    # Setup output paths
    task_id = task["id"]
    out_dir = os.path.join(OUTPUT_DIR, str(task_id))
    os.makedirs(out_dir, exist_ok=True)

    src_dir = os.path.join(VERIFIED_DIR, "spreadsheet", str(task_id))
    init_src = os.path.join(src_dir, f"1_{task_id}_init.xlsx")
    golden_src = os.path.join(src_dir, f"1_{task_id}_golden.xlsx")
    init_dst = os.path.join(out_dir, f"1_{task_id}_init.xlsx")
    golden_dst = os.path.join(out_dir, f"1_{task_id}_golden.xlsx")

    shutil.copy2(init_src, init_dst)
    shutil.copy2(golden_src, golden_dst)

    # 4a. Bake formulas in golden
    _bake_formulas(golden_dst)

    # 4b. Apply renames
    rename = spec.get("rename", {})
    _apply_rename(init_dst, rename, schema)
    _apply_rename(golden_dst, rename, schema)
    for old, new in rename.get("sheet_rename_map", {}).items():
        coord_map.add_sheet_rename(old, new)

    # 4c. Apply layout shift
    layout = spec.get("layout", {})
    _apply_layout(init_dst, layout, coord_map)
    _apply_layout(golden_dst, layout, coord_map)

    # 4d. Apply distractor
    distractor = spec.get("distractor", {})
    _apply_distractor(init_dst, distractor, coord_map)
    _apply_distractor(golden_dst, distractor, coord_map)

    # 4e. Update answer position
    new_ap = coord_map.map_answer_position(task["answer_position"])

    return coord_map, {
        "init_path": init_dst,
        "golden_path": golden_dst,
        "answer_position": new_ap,
        "out_dir": out_dir,
    }


def stage_4_5_check_layout_impact(instruction, spec, perturbed_paths, risk_flags):
    """LLM #3.5: Guardrail — check layout doesn't change difficulty."""
    layout = spec.get("layout", {})
    if layout.get("type") == "skip":
        return {"safe": True, "issues": []}

    preview = _gen_preview(perturbed_paths["init_path"])
    prompt = LAYOUT_IMPACT_PROMPT.format(
        instruction=instruction[:1000],
        layout_type=layout.get("type", "skip"),
        target_sheets=layout.get("target_sheets", []),
        perturbed_preview=preview,
        risk_flags=risk_flags,
    )
    return call_llm_json(prompt, model=LLM_MODEL)


def stage_5_rewrite_instruction(instruction, spec, coord_map):
    """LLM #2: Rewrite instruction for perturbation. Returns (instruction_text, issues).
    Uses structured output + deterministic post-checks."""
    rename = spec.get("rename", {})
    layout = spec.get("layout", {})
    distractor = spec.get("distractor", {})

    row_shift_desc = "No row shift"
    if layout.get("type") in ("title_row_insert", "blank_row_insert"):
        row_shift_desc = "All rows shifted down by 1 (row 1 becomes row 2, etc.)"

    distractor_desc = "None"
    if distractor.get("type") == "column":
        distractor_desc = f"Added distractor column '{distractor.get('name', '')}'"

    paraphrase = spec.get("instruction_paraphrase", False)
    paraphrase_instruction = ""
    if paraphrase:
        paraphrase_instruction = "9. Also paraphrase the instruction (different wording, same meaning)."

    # Inject feedback from previous quality check failures
    feedback_section = ""
    rewrite_feedback = spec.get("_rewrite_feedback", [])
    if rewrite_feedback:
        feedback_section = "\n\n## PREVIOUS REWRITE REJECTED — fix these issues:\n"
        for i, fb in enumerate(rewrite_feedback):
            feedback_section += f"- Attempt {i+1}: {fb}\n"
        feedback_section += "\nDo NOT repeat these mistakes."

    prompt = REWRITE_INSTRUCTION_PROMPT.format(
        column_rename_map=json.dumps(rename.get("column_rename_map", {})),
        sheet_rename_map=json.dumps(rename.get("sheet_rename_map", {})),
        row_shift_description=row_shift_desc,
        distractor_description=distractor_desc,
        paraphrase=paraphrase,
        instruction=instruction,
        paraphrase_instruction=paraphrase_instruction,
    )
    prompt += feedback_section

    result = call_llm_json(prompt, system="You are a precise text rewriter. Respond only with valid JSON.",
                           model=LLM_MODEL, max_tokens=4096)

    final_instruction = result.get("final_instruction", "").strip()

    # Fallback: if LLM returned plain text instead of JSON
    if not final_instruction:
        # Try raw text response
        raw = call_llm(prompt, system="You are a precise text rewriter.", model=LLM_MODEL,
                       max_tokens=4096).strip()
        final_instruction = raw

    # Deterministic post-fix: apply any missed renames that the LLM skipped
    final_instruction = _apply_missed_renames(instruction, final_instruction, spec)

    # Deterministic post-checks on the rewritten instruction
    post_issues = _check_rewrite_deterministic(instruction, final_instruction, spec)

    return final_instruction, post_issues


def stage_5_5_check_instruction_quality(original, perturbed, spec):
    """LLM #4.5: Guardrail — check instruction quality."""
    rename = spec.get("rename", {})
    prompt = INSTRUCTION_QUALITY_PROMPT.format(
        original_instruction=original[:1000],
        perturbed_instruction=perturbed[:1000],
        column_rename_map=json.dumps(rename.get("column_rename_map", {})),
        sheet_rename_map=json.dumps(rename.get("sheet_rename_map", {})),
    )
    return call_llm_json(prompt, model=LLM_MODEL)


def stage_6_validate_code(perturbed_paths, spec, schema, coord_map, original_ap,
                          original_golden_path=None):
    """Code: Deterministic validation checks."""
    issues = []

    # 6a. Golden self-check
    golden_ok = _golden_self_check(perturbed_paths["golden_path"],
                                    perturbed_paths["answer_position"])
    if not golden_ok:
        issues.append({"severity": "P0", "criterion": "golden_self_check",
                        "description": "Golden vs golden evaluation failed"})

    # 6b. Namespace validity
    issues.extend(_check_namespace(spec, schema))

    # 6c. Old name leftovers (checked after instruction rewrite)
    # Will be called separately with the rewritten instruction

    # 6d. Formula baking
    issues.extend(_check_formula_baking(perturbed_paths["golden_path"]))

    # 6e. Range shape preservation
    issues.extend(_check_range_shape(original_ap, perturbed_paths["answer_position"]))

    # 6f. Golden content equivalence: original golden[original_ap] == perturbed golden[perturbed_ap]
    if original_golden_path:
        issues.extend(_check_golden_content_equivalence(
            original_golden_path, original_ap,
            perturbed_paths["golden_path"], perturbed_paths["answer_position"],
            spec=spec))

    return issues


def stage_6_check_old_name_leftovers(perturbed_instruction, spec):
    """Code: Check old names don't appear in perturbed instruction."""
    issues = []
    rename = spec.get("rename", {})
    for old in list(rename.get("column_rename_map", {}).keys()) + \
               list(rename.get("sheet_rename_map", {}).keys()):
        if len(old) > 3 and old in perturbed_instruction:
            new = rename.get("column_rename_map", {}).get(old,
                  rename.get("sheet_rename_map", {}).get(old, ""))
            if old not in new:
                issues.append({"severity": "P0", "criterion": "old_name_leftover",
                               "description": f"Old name '{old}' still in instruction"})
    return issues


def stage_6_5_final_sanity(perturbed_instruction, perturbed_paths, spec,
                            original_ap, code_issues):
    """LLM #5.5: Guardrail — final sanity check."""
    preview = _gen_preview(perturbed_paths["init_path"])
    rename = spec.get("rename", {})
    prompt = FINAL_SANITY_PROMPT.format(
        perturbed_instruction=perturbed_instruction[:1000],
        perturbed_preview=preview,
        original_ap=original_ap,
        perturbed_ap=perturbed_paths["answer_position"],
        code_validation_summary=json.dumps(code_issues, indent=2)[:500],
        anchor_types=spec.get("anchor_types_used", []),
        rename_summary=json.dumps({
            "columns": rename.get("column_rename_map", {}),
            "sheets": rename.get("sheet_rename_map", {}),
        }),
        layout_summary=json.dumps(spec.get("layout", {})),
    )
    return call_llm_json(prompt, model=LLM_MODEL)


# ============================================================
# Main pipeline: per-task with feedback loop
# ============================================================

def process_one_task(task, family_guides):
    """Process one task through the full pipeline with feedback loop."""
    task_id = str(task["id"])
    instruction = task["instruction"]
    answer_position = task["answer_position"]

    try:
        # Stage 1: Schema extract
        init_path = os.path.join(VERIFIED_DIR, "spreadsheet", task_id, f"1_{task_id}_init.xlsx")
        golden_path = os.path.join(VERIFIED_DIR, "spreadsheet", task_id, f"1_{task_id}_golden.xlsx")

        if not os.path.exists(init_path):
            return {"task_id": task_id, "status": "excluded", "reason": "init file not found"}

        schema = extract_schema(task_id, init_path, golden_path, instruction, answer_position)

        # Stage 2: Feasibility
        feasibility = check_feasibility(schema)
        if not feasibility["feasible"]:
            return {"task_id": task_id, "status": "excluded", "reason": feasibility["reason"]}

        safe_anchors = feasibility["safe_anchors"]
        formula_policy = feasibility["formula_policy"]

        # LLM #0: Family classification
        family, risk_flags = stage_0_classify_family(instruction, schema)
        family_guide = family_guides.get(family, "No specific guide for this family.")

        # Feedback loop
        error_history = []

        for attempt in range(MAX_RETRIES + 1):
            # LLM #1: Design perturbation
            spec = stage_3_design_perturbation(
                schema, instruction, answer_position,
                family, family_guide, safe_anchors, risk_flags, error_history)

            # LLM #1.5: Rename safety guardrail
            rename_check = stage_3_5_check_rename_safety(spec, instruction, schema)
            if not rename_check.get("approved", True):
                error_history.append({
                    "stage": "rename_safety", "attempt": attempt,
                    "issues": rename_check.get("issues", []),
                    "spec": _summarize_spec(spec),
                })
                continue

            # Stage 4: Apply perturbation
            coord_map, perturbed_paths = stage_4_apply_perturbation(
                task, spec, schema, formula_policy)

            # LLM #2: Rewrite instruction (with inner retry on quality failure)
            rewrite_ok = False
            new_instruction = None
            for rewrite_attempt in range(3):
                new_instruction, rewrite_post_issues = stage_5_rewrite_instruction(
                    instruction, spec, coord_map)

                # Deterministic post-check failures are immediate P0
                rewrite_p0 = [i for i in rewrite_post_issues if i["severity"] == "P0"]
                if rewrite_p0:
                    if rewrite_attempt < 2:
                        spec.setdefault("_rewrite_feedback", [])
                        spec["_rewrite_feedback"].append(
                            "; ".join(i["description"] for i in rewrite_p0))
                    continue

                # LLM #4.5: Instruction quality guardrail
                quality_check = stage_5_5_check_instruction_quality(instruction, new_instruction, spec)
                if quality_check.get("pass", True):
                    rewrite_ok = True
                    break

                # Inner retry: inject quality issues into the rewrite prompt
                quality_issues = quality_check.get("issues", [])
                issue_text = "; ".join(
                    f"{i.get('type','')}: {i.get('evidence','')}" for i in quality_issues[:3])
                if rewrite_attempt < 2:
                    spec.setdefault("_rewrite_feedback", [])
                    spec["_rewrite_feedback"].append(issue_text)

            if not rewrite_ok:
                # Combine deterministic + LLM issues for error history
                all_rewrite_issues = rewrite_post_issues + quality_check.get("issues", [])
                error_history.append({
                    "stage": "instruction_quality", "attempt": attempt,
                    "issues": all_rewrite_issues[:5],
                    "spec": _summarize_spec(spec),
                })
                spec.pop("_rewrite_feedback", None)
                continue

            spec.pop("_rewrite_feedback", None)

            # LLM #3.5: Layout impact guardrail (AFTER rewrite, so it sees updated refs)
            layout_check = stage_4_5_check_layout_impact(
                new_instruction, spec, perturbed_paths, risk_flags)
            if not layout_check.get("safe", True):
                p0s = [i for i in layout_check.get("issues", []) if i.get("severity") == "P0"]
                if p0s:
                    error_history.append({
                        "stage": "layout_impact", "attempt": attempt,
                        "issues": p0s, "spec": _summarize_spec(spec),
                    })
                    continue

            # Stage 6: Code validation
            code_issues = stage_6_validate_code(
                perturbed_paths, spec, schema, coord_map, answer_position,
                original_golden_path=golden_path)
            code_issues.extend(stage_6_check_old_name_leftovers(new_instruction, spec))

            p0_code = [i for i in code_issues if i["severity"] == "P0"]
            if p0_code:
                error_history.append({
                    "stage": "code_validation", "attempt": attempt,
                    "issues": p0_code, "spec": _summarize_spec(spec),
                })
                continue

            # LLM #5.5: Final sanity guardrail
            sanity = stage_6_5_final_sanity(
                new_instruction, perturbed_paths, spec, answer_position, code_issues)
            if sanity.get("verdict") == "fail":
                error_history.append({
                    "stage": "final_sanity", "attempt": attempt,
                    "issues": sanity.get("issues", []),
                    "spec": _summarize_spec(spec),
                })
                continue

            # All passed — save outputs
            _save_outputs(task_id, new_instruction, perturbed_paths, spec, coord_map)

            return {
                "task_id": task_id,
                "status": "pass",
                "family": family,
                "anchor_types_used": spec.get("anchor_types_used", []),
                "spec": spec,
                "coord_map": coord_map.to_dict(),
                "answer_position_before": answer_position,
                "answer_position_after": perturbed_paths["answer_position"],
                "hashes": _compute_hashes(task_id, perturbed_paths),
                "validation": {
                    "code_p0": len(p0_code),
                    "code_p1": len([i for i in code_issues if i["severity"] == "P1"]),
                    "golden_self_check": True,
                    "formula_policy": formula_policy,
                },
                "attempt": attempt,
                "error_history": error_history,
            }

        # Exhausted retries
        return {
            "task_id": task_id,
            "status": "fail",
            "error_history": error_history,
        }

    except Exception as e:
        return {
            "task_id": task_id,
            "status": "error",
            "error": str(e),
            "traceback": traceback.format_exc()[:500],
        }


# ============================================================
# Orchestrator
# ============================================================

def run_pipeline(max_workers=32):
    """Run the full pipeline on all 400 tasks."""
    # Load dataset
    with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
        dataset = json.load(f)

    # Load family guides
    family_guides = _load_family_guides()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Process all tasks in parallel
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {}
        for task in dataset:
            future = pool.submit(process_one_task, task, family_guides)
            futures[future] = str(task["id"])

        for future in as_completed(futures):
            tid = futures[future]
            try:
                result = future.result()
            except Exception as e:
                result = {"task_id": tid, "status": "error", "error": str(e)}

            results[tid] = result
            status = result.get("status", "?")
            n_done = len(results)
            print(f"[{n_done}/{len(dataset)}] {tid}: {status}")

    # Save all manifests
    manifest_path = os.path.join(OUTPUT_DIR, "all_manifests.json")
    with open(manifest_path, "w") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    # Summary
    statuses = Counter(r["status"] for r in results.values())
    print(f"\n{'='*60}")
    print(f"PIPELINE COMPLETE: {len(results)} tasks")
    for s, c in sorted(statuses.items()):
        print(f"  {s}: {c}")
    print(f"Results saved to {manifest_path}")

    return results


# ============================================================
# Helper stubs (to be filled)
# ============================================================

def _apply_missed_renames(original, rewritten, spec):
    """Deterministic post-fix: apply column/sheet renames that the LLM missed.
    Only replaces if the old name appears in the rewritten text in the same context
    as it appeared in the original (to avoid replacing common English words incorrectly)."""
    rename = spec.get("rename", {})

    for old, new in rename.get("column_rename_map", {}).items():
        if len(old) <= 2:
            continue  # Too short, high false-positive risk
        if old in rewritten and old not in new:
            # Verify it appeared in the original too (not something the LLM added)
            if old in original:
                rewritten = rewritten.replace(old, new)

    for old, new in rename.get("sheet_rename_map", {}).items():
        if len(old) <= 1:
            continue
        if old in rewritten and old not in new:
            if old in original:
                rewritten = rewritten.replace(old, new)

    return rewritten


def _check_rewrite_deterministic(original, rewritten, spec):
    """Deterministic post-checks on rewritten instruction.
    Catches leakage, length inflation, and unauthorized additions without LLM."""
    issues = []

    # 1. Forbidden leakage phrases
    LEAKAGE_PHRASES = [
        "row shifted", "rows shifted", "shifted down", "shifted up",
        "distractor", "ignore the column", "ignore inserted",
        "perturbation", "perturbed", "renamed", "was renamed",
        "column was added", "extra column", "additional column",
        "blank row", "title row", "inserted row", "inserted column",
        "structural change", "layout change",
    ]
    rewritten_lower = rewritten.lower()
    for phrase in LEAKAGE_PHRASES:
        if phrase in rewritten_lower and phrase not in original.lower():
            issues.append({
                "severity": "P0", "criterion": "rewrite_leakage",
                "description": f"Leaked perturbation detail: '{phrase}' not in original",
            })

    # 2. Length inflation: rewritten should not be >35% longer than original
    orig_len = len(original)
    rew_len = len(rewritten)
    if orig_len > 0 and rew_len > orig_len * 1.35:
        issues.append({
            "severity": "P1", "criterion": "rewrite_length_inflation",
            "description": f"Rewritten is {rew_len/orig_len:.0%} of original length ({rew_len} vs {orig_len} chars)",
        })

    # 3. Check all renames were applied
    rename = spec.get("rename", {})
    for old, new in rename.get("column_rename_map", {}).items():
        if len(old) > 3 and old in rewritten and old not in new:
            issues.append({
                "severity": "P0", "criterion": "rewrite_incomplete_rename",
                "description": f"Old column name '{old}' still in rewritten instruction",
            })
    for old, new in rename.get("sheet_rename_map", {}).items():
        if len(old) > 2 and old in rewritten and old not in new:
            issues.append({
                "severity": "P0", "criterion": "rewrite_incomplete_rename",
                "description": f"Old sheet name '{old}' still in rewritten instruction",
            })

    return issues


def _format_schema_summary(schema):
    """Format schema for LLM prompt."""
    lines = []
    for sn, info in schema["sheets"].items():
        headers = [h["name"] for h in info.get("headers", [])][:6]
        lines.append(f"  Sheet '{sn}': {info['rows']}r x {info['cols']}c, "
                     f"headers={headers}, formulas={info['formula_count']}")
    return "\n".join(lines)


def _format_error_history(error_history):
    """Format error history for LLM prompt."""
    if not error_history:
        return ""
    lines = ["## Previous Attempts (FAILED — avoid these mistakes)"]
    for err in error_history:
        lines.append(f"\n### Attempt {err['attempt']+1} — failed at {err['stage']}")
        for issue in err.get("issues", []):
            desc = issue.get("description", issue.get("problem", ""))
            lines.append(f"- {desc}")
        spec = err.get("spec", {})
        if spec:
            lines.append(f"- Spec used: {json.dumps(spec)[:200]}")
    return "\n".join(lines)


def _summarize_spec(spec):
    """Compact spec summary for error history."""
    return {
        "anchor_types": spec.get("anchor_types_used", []),
        "col_renames": list(spec.get("rename", {}).get("column_rename_map", {}).keys()),
        "sheet_renames": list(spec.get("rename", {}).get("sheet_rename_map", {}).keys()),
        "layout_type": spec.get("layout", {}).get("type", "skip"),
    }


def _load_family_guides():
    """Load family-level perturbation guides, split by ## Family headers."""
    guides = {}
    guide_path = "scratch/family-perturbation-guides.md"
    if not os.path.exists(guide_path):
        return guides

    with open(guide_path) as f:
        content = f.read()

    # Map section headers to family keys
    family_header_map = {
        "Family A: Formula Write-back": "formula_writeback",
        "Family B: Lookup / Match / Merge": "lookup_match_merge",
        "Family C: Conditional Partition / Filtering": "conditional_partition",
        "Family D: Grouped Aggregation / Summary Write-back": "grouped_aggregation",
        "Family E: Structural Edit with Reference Preservation": "structural_edit",
    }

    # Split by ## headers
    import re
    sections = re.split(r'\n(?=## Family)', content)
    for section in sections:
        for header_text, family_key in family_header_map.items():
            if header_text in section:
                guides[family_key] = section.strip()[:2000]
                break

    return guides


def _bake_formulas(golden_path):
    """Bake formula cells in golden to static values.
    Reads the file with data_only=True to get cached values, then replaces
    formula cells in the original with those cached values."""
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula

    wb = openpyxl.load_workbook(golden_path)
    wb_data = openpyxl.load_workbook(golden_path, data_only=True)

    baked = 0
    unbaked = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        ws_data = wb_data[sn]
        for row in ws.iter_rows():
            for cell in row:
                is_formula = False
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    is_formula = True
                elif isinstance(cell.value, ArrayFormula):
                    is_formula = True

                if is_formula:
                    cached = ws_data[cell.coordinate].value
                    if cached is not None:
                        cell.value = cached
                        baked += 1
                    else:
                        unbaked += 1

    wb.save(golden_path)
    wb.close()
    wb_data.close()
    return {"baked": baked, "unbaked": unbaked}


def _apply_rename(file_path, rename_spec, schema):
    """Apply rename operations to a workbook file.
    Handles: sheet renames, header renames, data value renames, formula sheet-ref updates."""
    import openpyxl

    col_map = rename_spec.get("column_rename_map", {})
    sheet_map = rename_spec.get("sheet_rename_map", {})
    data_renames = rename_spec.get("data_value_renames", [])

    if not col_map and not sheet_map and not data_renames:
        return

    wb = openpyxl.load_workbook(file_path)

    # A. Sheet rename
    for old_name, new_name in sheet_map.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name

    # B. Header rename — only on the detected header row
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Find the schema entry (may be under old or new name)
        schema_entry = None
        for orig_sn, si in schema.get("sheets", {}).items():
            actual_name = sheet_map.get(orig_sn, orig_sn)
            if actual_name == sn:
                schema_entry = si
                break
        if not schema_entry:
            continue

        hdr_row = schema_entry.get("header_row")
        if hdr_row is None:
            continue

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=hdr_row, column=col_idx)
            if cell.value and str(cell.value) in col_map:
                cell.value = col_map[str(cell.value)]

    # C. Data value rename
    for dr in data_renames:
        find_text = dr.get("find", "")
        replace_text = dr.get("replace", "")
        target_sheet = dr.get("sheet")
        for sn in wb.sheetnames:
            if target_sheet and sn != target_sheet:
                # Also check if target_sheet was renamed
                renamed = sheet_map.get(target_sheet, target_sheet)
                if sn != renamed:
                    continue
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and find_text in cell.value:
                        cell.value = cell.value.replace(find_text, replace_text)

    # D. Formula sheet-ref update
    if sheet_map:
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        original = cell.value
                        for old_sn, new_sn in sheet_map.items():
                            cell.value = cell.value.replace(f"'{old_sn}'!", f"'{new_sn}'!")
                            cell.value = cell.value.replace(f"{old_sn}!", f"{new_sn}!")

    wb.save(file_path)
    wb.close()


def _apply_layout(file_path, layout_spec, coord_map):
    """Apply layout shift to a workbook file.
    Supports: title_row_insert, blank_row_insert.
    Updates coord_map with row shifts."""
    import openpyxl

    layout_type = layout_spec.get("type", "skip")
    if layout_type == "skip":
        return

    target_sheets = layout_spec.get("target_sheets", [])
    params = layout_spec.get("params", {})

    wb = openpyxl.load_workbook(file_path)

    for sn in wb.sheetnames:
        if target_sheets and sn not in target_sheets:
            continue

        ws = wb[sn]

        if layout_type == "title_row_insert":
            title_text = params.get("title_text", "Report")
            ws.insert_rows(1, amount=1)
            ws.cell(row=1, column=1, value=title_text)
            coord_map.add_row_shift(sn, 1)

        elif layout_type == "blank_row_insert":
            ws.insert_rows(1, amount=1)
            # Row 1 stays blank
            coord_map.add_row_shift(sn, 1)

    wb.save(file_path)
    wb.close()


def _apply_distractor(file_path, distractor_spec, coord_map):
    """Apply distractor column/sheet to a workbook file.
    Inserts a column with a plausible-sounding name and random-ish data."""
    import openpyxl
    import random

    dist_type = distractor_spec.get("type", "none")
    if dist_type == "none" or not dist_type:
        return

    wb = openpyxl.load_workbook(file_path)

    if dist_type == "column":
        col_name = distractor_spec.get("name", "Notes")
        target_sheet = distractor_spec.get("target_sheet")
        position = distractor_spec.get("position", "after_last_column")

        for sn in wb.sheetnames:
            if target_sheet and sn != target_sheet:
                continue
            ws = wb[sn]

            if position == "after_last_column":
                insert_col = ws.max_column + 1
            else:
                # Try to parse a column index
                try:
                    insert_col = int(position)
                except (ValueError, TypeError):
                    insert_col = ws.max_column + 1

            # If inserting in the middle, shift existing columns
            if insert_col <= ws.max_column:
                ws.insert_cols(insert_col, amount=1)
                coord_map.add_col_insert(sn, insert_col, 1)

            # Write header
            header_row = 1
            # Try to detect header row from existing data
            for r in range(1, min(5, ws.max_row + 1)):
                cell_val = ws.cell(row=r, column=1).value
                if cell_val and isinstance(cell_val, str):
                    header_row = r
                    break

            ws.cell(row=header_row, column=insert_col, value=col_name)

            # Fill with empty/placeholder values (distractor should be ignorable)
            fill_values = ["", None, "N/A", "-"]
            for row_idx in range(header_row + 1, min(ws.max_row + 1, header_row + 200)):
                ws.cell(row=row_idx, column=insert_col,
                        value=random.choice(fill_values))

    elif dist_type == "sheet":
        sheet_name = distractor_spec.get("name", "Notes")
        ws_new = wb.create_sheet(title=sheet_name)
        ws_new.cell(row=1, column=1, value="This sheet is intentionally left mostly blank.")

    wb.save(file_path)
    wb.close()


def _gen_preview(file_path, max_rows=6):
    """Generate text preview of spreadsheet (first N rows per sheet)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = ""
        for sn in wb.sheetnames:
            ws = wb[sn]
            result += f"Sheet '{sn}':\n"
            for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
                vals = [str(ws.cell(row_idx, c).value or "") for c in range(1, min(12, ws.max_column + 1))]
                result += f"  Row {row_idx}: {vals}\n"
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


def _golden_self_check(golden_path, answer_position):
    """Check golden vs golden passes evaluation.
    Loads the golden file, evaluates it against itself at the answer_position."""
    import openpyxl
    import re
    import datetime

    if not os.path.exists(golden_path):
        return False

    def transform_value(v):
        if isinstance(v, (int, float)):
            v = round(float(v), 2)
        elif isinstance(v, datetime.time):
            v = str(v)[:-3]
        elif isinstance(v, datetime.datetime):
            excel_start = datetime.datetime(1899, 12, 30)
            delta = v - excel_start
            v = round(delta.days + delta.seconds / 86400.0, 0)
        elif isinstance(v, str):
            try:
                v = round(float(v), 2)
            except ValueError:
                pass
        return v

    def compare_cell_value(v1, v2):
        v1 = transform_value(v1)
        v2 = transform_value(v2)
        if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
            return True
        if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
            return True
        if type(v1) != type(v2):
            return False
        return v1 == v2

    def col_name2num(name):
        num = 0
        for c in name:
            num = num * 26 + (ord(c) - ord("A") + 1)
        return num

    def col_num2name(n):
        name = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            name = chr(65 + remainder) + name
        return name

    def generate_cell_names(range_str):
        if ":" not in range_str:
            return [range_str]
        start, end = range_str.split(":")
        sc, sr = "", ""
        for c in start:
            if c.isdigit():
                sr += c
            else:
                sc += c
        ec, er = "", ""
        for c in end:
            if c.isdigit():
                er += c
            else:
                ec += c
        cols = [col_num2name(i) for i in range(col_name2num(sc), col_name2num(ec) + 1)]
        return [f"{col}{row}" for col in cols for row in range(int(sr), int(er) + 1)]

    try:
        wb = openpyxl.load_workbook(golden_path, data_only=True)

        # Normalize answer_position
        ap = re.sub(r"'([^'!]+)!'", r"'\1'!", answer_position)
        ap = ap.rstrip("'").rstrip('"')
        parts = re.split(r",(?='|[A-Z])", ap)

        for part in parts:
            part = part.strip()
            if "!" in part:
                sheet_name, cell_range = part.rsplit("!", 1)
                sheet_name = sheet_name.strip("'").strip('"')
            else:
                sheet_name = wb.sheetnames[0]
                cell_range = part

            cell_range = cell_range.strip("'").strip('"')

            if sheet_name not in wb.sheetnames:
                wb.close()
                return False

            ws = wb[sheet_name]
            for cell_name in generate_cell_names(cell_range):
                val = ws[cell_name].value
                if not compare_cell_value(val, val):
                    wb.close()
                    return False

                # Also check that the cell is not empty (golden should have values)
                # Allow empty if the task legitimately clears cells
                # For self-check, just verify the file is readable at these positions

        wb.close()
        return True
    except Exception:
        return False


def _check_namespace(spec, schema):
    """Check rename namespace validity: no collisions, no illegal chars, no cell-ref names."""
    import re
    issues = []
    rename = spec.get("rename", {})
    col_map = rename.get("column_rename_map", {})
    sheet_map = rename.get("sheet_rename_map", {})

    EXCEL_ILLEGAL_SHEET_CHARS = set("\\/*?:[]")
    EXCEL_MAX_SHEET_NAME_LEN = 31

    # Check sheet name legality
    for old, new in sheet_map.items():
        if len(new) > EXCEL_MAX_SHEET_NAME_LEN:
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Sheet name '{new}' exceeds Excel 31-char limit ({len(new)} chars)",
            })
        if any(c in new for c in EXCEL_ILLEGAL_SHEET_CHARS):
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Sheet name '{new}' contains illegal Excel characters",
            })
        if re.match(r'^[A-Z]{1,3}\d+$', new):
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Sheet name '{new}' looks like a cell reference",
            })

    # Check sheet name uniqueness after rename
    sheet_names = list(schema.get("sheet_names", []))
    renamed_sheets = [sheet_map.get(sn, sn) for sn in sheet_names]
    if len(renamed_sheets) != len(set(renamed_sheets)):
        dupes = [s for s in renamed_sheets if renamed_sheets.count(s) > 1]
        issues.append({
            "severity": "P0", "criterion": "rename_namespace",
            "description": f"Sheet name collision after rename: {set(dupes)}",
        })

    # Check header uniqueness after rename within each sheet
    for sn, info in schema.get("sheets", {}).items():
        headers = [h["name"] for h in info.get("headers", [])]
        renamed_headers = [col_map.get(h, h) for h in headers]
        if len(renamed_headers) != len(set(renamed_headers)):
            dupes = [h for h in renamed_headers if renamed_headers.count(h) > 1]
            issues.append({
                "severity": "P0", "criterion": "rename_namespace",
                "description": f"Header collision in sheet '{sn}' after rename: {set(dupes)}",
            })

    # Check that rename doesn't map two different old names to the same new name
    new_col_names = list(col_map.values())
    if len(new_col_names) != len(set(new_col_names)):
        dupes = [n for n in new_col_names if new_col_names.count(n) > 1]
        issues.append({
            "severity": "P0", "criterion": "rename_namespace",
            "description": f"Column rename maps multiple old names to same new name: {set(dupes)}",
        })

    return issues


def _check_formula_baking(golden_path):
    """Check golden has no remaining formulas after baking."""
    import openpyxl
    issues = []

    if not os.path.exists(golden_path):
        issues.append({
            "severity": "P0", "criterion": "formula_baking",
            "description": "Golden file not found",
        })
        return issues

    try:
        wb = openpyxl.load_workbook(golden_path)
        formula_count = 0
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb.close()

        if formula_count > 0:
            issues.append({
                "severity": "P0", "criterion": "formula_baking",
                "description": f"Golden file still has {formula_count} formula cells (should be 0 after baking)",
            })
    except Exception as e:
        issues.append({
            "severity": "P0", "criterion": "formula_baking",
            "description": f"Error reading golden file: {str(e)[:100]}",
        })

    return issues


def _check_range_shape(original_ap, perturbed_ap):
    """Check answer range shape is preserved after perturbation."""
    import re
    issues = []

    def parse_range_shape(ap):
        """Extract (rows, cols) from ranges like B3:C68."""
        ranges = re.findall(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', ap)
        shapes = []
        for c1, r1, c2, r2 in ranges:
            rows = int(r2) - int(r1) + 1
            cols = ord(c2[-1]) - ord(c1[-1]) + 1  # simple single-letter
            # Handle multi-letter columns
            def col_to_num(c):
                n = 0
                for ch in c:
                    n = n * 26 + (ord(ch) - ord('A') + 1)
                return n
            cols = col_to_num(c2) - col_to_num(c1) + 1
            shapes.append((rows, cols))
        return shapes

    orig_shapes = parse_range_shape(original_ap)
    pert_shapes = parse_range_shape(perturbed_ap)

    if orig_shapes and pert_shapes and orig_shapes != pert_shapes:
        issues.append({
            "severity": "P0", "criterion": "range_shape_preservation",
            "description": f"Range shape changed. Orig: {orig_shapes}, Pert: {pert_shapes}",
        })

    return issues


def _check_golden_content_equivalence(orig_golden_path, orig_ap, pert_golden_path, pert_ap,
                                       spec=None):
    """Check that cell values in original golden[orig_ap] match perturbed golden[pert_ap].
    This catches AP offset bugs, data loss, and bake errors in one shot."""
    import openpyxl
    import datetime
    issues = []

    def transform_value(v):
        if isinstance(v, (int, float)):
            return round(float(v), 2)
        elif isinstance(v, datetime.time):
            return str(v)[:-3]
        elif isinstance(v, datetime.datetime):
            excel_start = datetime.datetime(1899, 12, 30)
            delta = v - excel_start
            return round(delta.days + delta.seconds / 86400.0, 0)
        elif isinstance(v, str):
            try:
                return round(float(v), 2)
            except ValueError:
                return v
        return v

    def col_name2num(name):
        num = 0
        for c in name:
            num = num * 26 + (ord(c) - ord("A") + 1)
        return num

    def col_num2name(n):
        name = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            name = chr(65 + remainder) + name
        return name

    def parse_ap(ap):
        """Parse answer position into list of (sheet_or_None, cell_range) tuples."""
        # Normalize only the specific malformed case: word'! -> 'word'!
        # Must NOT touch already-quoted formats like 'Sheet1'!
        import re as _re
        # Only fix: bare word followed by '! at the very start of a part
        # We do this per-part after splitting

        ap = ap.rstrip("'").rstrip('"')
        # Split by comma respecting quotes
        parts = []
        current, in_q = "", False
        for ch in ap:
            if ch == "'":
                in_q = not in_q
            if ch == "," and not in_q:
                parts.append(current.strip())
                current = ""
            else:
                current += ch
        if current.strip():
            parts.append(current.strip())

        result = []
        for part in parts:
            # Normalize per-part: word'!X -> 'word'!X (only if no leading quote)
            part = _re.sub(r"^(\w+)'!", r"'\1'!", part)
            # Normalize: 'Sheet1!'X -> 'Sheet1'!X
            part = _re.sub(r"'([^'!]+)!'", r"'\1'!", part)

            if "!" in part:
                sheet_part, cell_range = part.rsplit("!", 1)
                sheet = sheet_part.strip("'").strip('"')
            else:
                sheet = None
                cell_range = part
            result.append((sheet, cell_range.strip("'").strip('"')))
        return result

    def get_range_values(wb, sheet, cell_range):
        """Extract values from a range like A14:I137."""
        if sheet:
            if sheet not in wb.sheetnames:
                return None
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]

        if ":" not in cell_range:
            return [transform_value(ws[cell_range].value)]

        start, end = cell_range.split(":")
        sc, sr, ec, er = "", "", "", ""
        for c in start:
            if c.isdigit():
                sr += c
            else:
                sc += c
        for c in end:
            if c.isdigit():
                er += c
            else:
                ec += c

        values = []
        for col_idx in range(col_name2num(sc), col_name2num(ec) + 1):
            for row_idx in range(int(sr), int(er) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                values.append(transform_value(val))
        return values

    try:
        wb_orig = openpyxl.load_workbook(orig_golden_path, data_only=True)
        wb_pert = openpyxl.load_workbook(pert_golden_path, data_only=True)

        orig_parts = parse_ap(orig_ap)
        pert_parts = parse_ap(pert_ap)

        if len(orig_parts) != len(pert_parts):
            issues.append({
                "severity": "P0", "criterion": "golden_content_equivalence",
                "description": f"AP part count mismatch: orig has {len(orig_parts)}, pert has {len(pert_parts)}",
            })
            wb_orig.close()
            wb_pert.close()
            return issues

        for (o_sheet, o_range), (p_sheet, p_range) in zip(orig_parts, pert_parts):
            orig_vals = get_range_values(wb_orig, o_sheet, o_range)
            pert_vals = get_range_values(wb_pert, p_sheet, p_range)

            if orig_vals is None or pert_vals is None:
                issues.append({
                    "severity": "P0", "criterion": "golden_content_equivalence",
                    "description": f"Could not read range: orig={o_sheet}!{o_range}, pert={p_sheet}!{p_range}",
                })
                continue

            if len(orig_vals) != len(pert_vals):
                issues.append({
                    "severity": "P0", "criterion": "golden_content_equivalence",
                    "description": f"Cell count mismatch: orig={len(orig_vals)}, pert={len(pert_vals)}",
                })
                continue

            # Build canonical rename map: apply approved renames to original values
            rename_map = {}
            if spec:
                rename = spec.get("rename", {})
                for old, new in rename.get("column_rename_map", {}).items():
                    rename_map[old] = new
                for old, new in rename.get("sheet_rename_map", {}).items():
                    rename_map[old] = new
                for dr in rename.get("data_value_renames", []):
                    rename_map[dr.get("find", "")] = dr.get("replace", "")

            def apply_canonical(val):
                """Apply approved renames to a value for comparison."""
                if isinstance(val, str):
                    for old, new in rename_map.items():
                        if old and old in val:
                            val = val.replace(old, new)
                return val

            mismatches = 0
            unexplained_mismatches = 0
            first_mismatch = None
            for i, (ov, pv) in enumerate(zip(orig_vals, pert_vals)):
                if ov == pv:
                    continue
                # Try canonical comparison (apply renames to original)
                canonical_ov = apply_canonical(ov) if isinstance(ov, str) else ov
                if canonical_ov == pv:
                    continue  # Explained by approved rename
                mismatches += 1
                unexplained_mismatches += 1
                if first_mismatch is None:
                    first_mismatch = (i, ov, pv)

            total = max(len(orig_vals), 1)

            # Dual threshold: stricter for small ranges
            if unexplained_mismatches == 0:
                pass  # All good
            elif total <= 50 and unexplained_mismatches > 0:
                # Small range: any unexplained mismatch is P0
                idx, ov, pv = first_mismatch
                issues.append({
                    "severity": "P0", "criterion": "golden_content_equivalence",
                    "description": (
                        f"{unexplained_mismatches}/{total} unexplained mismatches in small range. "
                        f"First at index {idx}: orig={ov!r}, pert={pv!r}."
                    ),
                })
            elif unexplained_mismatches / total > 0.10:
                # Large range: >10% unexplained is P0
                idx, ov, pv = first_mismatch
                issues.append({
                    "severity": "P0", "criterion": "golden_content_equivalence",
                    "description": (
                        f"{unexplained_mismatches}/{total} unexplained mismatches ({unexplained_mismatches/total:.0%}). "
                        f"First at index {idx}: orig={ov!r}, pert={pv!r}. "
                        f"AP may be offset or data corrupted."
                    ),
                })
            else:
                # Large range, ≤10% unexplained: P1 warning
                idx, ov, pv = first_mismatch
                issues.append({
                    "severity": "P1", "criterion": "golden_content_equivalence",
                    "description": (
                        f"{unexplained_mismatches}/{total} unexplained mismatches ({unexplained_mismatches/total:.0%}). "
                        f"First: orig={ov!r}, pert={pv!r}"
                    ),
                })

        wb_orig.close()
        wb_pert.close()
    except Exception as e:
        issues.append({
            "severity": "P0", "criterion": "golden_content_equivalence",
            "description": f"Error comparing golden files: {str(e)[:150]}",
        })

    return issues


def _save_outputs(task_id, instruction, perturbed_paths, spec, coord_map):
    """Save perturbed instruction and task_meta."""
    out_dir = perturbed_paths["out_dir"]

    with open(os.path.join(out_dir, "instruction_perturbed.txt"), "w") as f:
        f.write(instruction)

    task_meta = {
        "id": task_id,
        "instruction": instruction,
        "answer_position": perturbed_paths["answer_position"],
    }
    with open(os.path.join(out_dir, "task_meta.json"), "w") as f:
        json.dump(task_meta, f, indent=2, ensure_ascii=False)

    manifest = {
        "task_id": task_id,
        "spec": spec,
        "coord_map": coord_map.to_dict(),
    }
    with open(os.path.join(out_dir, "manifest.json"), "w") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)


def _compute_hashes(task_id, perturbed_paths):
    """Compute file hashes for manifest."""
    hashes = {}
    for key, path in [("perturbed_init", perturbed_paths["init_path"]),
                       ("perturbed_golden", perturbed_paths["golden_path"])]:
        if os.path.exists(path):
            with open(path, "rb") as f:
                hashes[key] = hashlib.sha256(f.read()).hexdigest()[:16]
    return hashes


# ============================================================
# CLI
# ============================================================

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--workers", type=int, default=32)
    parser.add_argument("--task_id", default=None, help="Process single task")
    parser.add_argument("--dry_run", action="store_true")
    args = parser.parse_args()

    if args.task_id:
        with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
            dataset = json.load(f)
        task = next(d for d in dataset if str(d["id"]) == args.task_id)
        family_guides = _load_family_guides()
        result = process_one_task(task, family_guides)
        print(json.dumps(result, indent=2))
    else:
        run_pipeline(max_workers=args.workers)
