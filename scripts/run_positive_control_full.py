#!/usr/bin/env python3
"""
Full-scale positive control: 77 tasks with natural rename + old-anchor distractor.
Uses actual successful code as hard_patch.
"""

import json
import os
import sys
import random
import shutil
import tempfile
import threading
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.path.insert(0, ".")
from scripts.run_pilot_experiment import (
    call_llm, extract_code, execute_code, evaluate_task, gen_spreadsheet_content, VERIFIED_DIR
)
from scripts.run_positive_control import (
    gen_natural_rename, gen_natural_sheet_rename, CROSS_DOMAIN_NAMES, CROSS_DOMAIN_SHEET_NAMES
)

OUTPUT_DIR = "data/positive_control_full"
RESULTS_DIR = "scratch/experiment_results"
LLM_MODEL = "us.anthropic.claude-3-5-sonnet-20241022-v2:0"


def generate_perturbed(tid, col_renames, sheet_renames):
    """Generate natural-renamed + distractor perturbed workbook."""
    src_dir = os.path.join(VERIFIED_DIR, "spreadsheet", tid)
    out_dir = os.path.join(OUTPUT_DIR, "perturbed", tid)
    os.makedirs(out_dir, exist_ok=True)

    init_src = os.path.join(src_dir, f"1_{tid}_init.xlsx")
    golden_src = os.path.join(src_dir, f"1_{tid}_golden.xlsx")

    # Get original headers
    wb = openpyxl.load_workbook(init_src)
    original_headers = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val:
                headers.append(str(val))
        original_headers[sn] = headers
    wb.close()

    # Generate natural renames
    random.seed(hash(tid) % 2**32)
    used_names = set()
    new_col_map = {}
    for old in col_renames.keys():
        new_col_map[old] = gen_natural_rename(old, used_names)
    new_sheet_map = {}
    for old in sheet_renames.keys():
        new_sheet_map[old] = gen_natural_sheet_rename(old, used_names)

    primary_distractor = list(col_renames.keys())[0] if col_renames else None

    # Apply to files
    for src, dst_name in [(init_src, f"1_{tid}_init.xlsx"), (golden_src, f"1_{tid}_golden.xlsx")]:
        wb = openpyxl.load_workbook(src)
        
        # Bake formulas in golden (replace formula cells with cached values)
        if "golden" in dst_name:
            from openpyxl.worksheet.formula import ArrayFormula
            wb_data = openpyxl.load_workbook(src, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                ws_data = wb_data[sn]
                for row in ws.iter_rows():
                    for cell in row:
                        is_formula = False
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            is_formula = True
                        elif isinstance(cell.value, ArrayFormula):
                            is_formula = True
                        if is_formula:
                            cached = ws_data[cell.coordinate].value
                            if cached is not None:
                                cell.value = cached
            wb_data.close()
        
        for old_sn, new_sn in new_sheet_map.items():
            if old_sn in wb.sheetnames:
                wb[old_sn].title = new_sn
        for sn in wb.sheetnames:
            ws = wb[sn]
            for c in range(1, ws.max_column + 1):
                val = ws.cell(1, c).value
                if val and str(val) in new_col_map:
                    ws.cell(1, c).value = new_col_map[str(val)]
            # Add distractor
            if primary_distractor:
                orig_sn = sn
                for old_s, new_s in new_sheet_map.items():
                    if new_s == sn:
                        orig_sn = old_s
                if primary_distractor in original_headers.get(orig_sn, []):
                    dist_col = ws.max_column + 1
                    ws.cell(1, dist_col).value = primary_distractor
                    for row in range(2, min(ws.max_row + 1, 200)):
                        ws.cell(row, dist_col).value = round(random.uniform(-100, 100), 2)
        wb.save(os.path.join(out_dir, dst_name))
        wb.close()

    # Update instruction + AP
    with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
        dataset = json.load(f)
    meta = next(d for d in dataset if str(d["id"]) == tid)

    instruction = meta["instruction"]
    for old, new in new_col_map.items():
        instruction = instruction.replace(old, new)
    for old, new in new_sheet_map.items():
        instruction = instruction.replace(old, new)

    ap = meta["answer_position"]
    for old, new in new_sheet_map.items():
        ap = ap.replace(old, new)

    task_meta = {"id": tid, "instruction": instruction, "answer_position": ap,
                 "instruction_type": meta.get("instruction_type", "Cell-Level Manipulation")}
    with open(os.path.join(out_dir, "task_meta.json"), "w") as f:
        json.dump(task_meta, f, indent=2)

    return {
        "old_anchors": list(new_col_map.keys()) + list(new_sheet_map.keys()),
        "new_anchors": list(new_col_map.values()) + list(new_sheet_map.values()),
        "col_map": new_col_map, "sheet_map": new_sheet_map,
    }


def run_one(tid, condition, patch_type, patch_content, manifest):
    if condition == "original":
        spread_dir = os.path.join(VERIFIED_DIR, "spreadsheet", tid)
        with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
            dataset = json.load(f)
        meta = next(d for d in dataset if str(d["id"]) == tid)
    else:
        spread_dir = os.path.join(OUTPUT_DIR, "perturbed", tid)
        with open(os.path.join(spread_dir, "task_meta.json")) as f:
            meta = json.load(f)

    input_file = f"1_{tid}_init.xlsx"
    golden_path = os.path.join(spread_dir, f"1_{tid}_golden.xlsx")
    input_path = os.path.join(spread_dir, input_file)
    if not os.path.exists(input_path):
        return {"task_id": tid, "condition": condition, "patch_type": patch_type,
                "pass": False, "code": "", "old_anchor_count": 0, "new_anchor_count": 0}

    with tempfile.TemporaryDirectory() as work_dir:
        shutil.copy2(input_path, os.path.join(work_dir, input_file))
        output_path = os.path.join(work_dir, f"1_{tid}_output.xlsx")
        preview = gen_spreadsheet_content(os.path.join(work_dir, input_file))

        artifact = f"{patch_content}\n\n" if patch_content else ""
        prompt = f"""You are a spreadsheet expert who can manipulate spreadsheets through Python code.

{artifact}You need to solve the given spreadsheet manipulation question:
### instruction
{meta['instruction']}

### spreadsheet_path
{os.path.join(work_dir, input_file)}

### spreadsheet_content
{preview}

### instruction_type
{meta.get('instruction_type', 'Cell-Level Manipulation')}

### answer_position
{meta['answer_position']}

### output_path
{output_path}

You should generate Python code for the final solution of the question."""

        try:
            response = call_llm(prompt, LLM_MODEL, None, None)
        except Exception as e:
            return {"task_id": tid, "condition": condition, "patch_type": patch_type,
                    "pass": False, "code": "", "old_anchor_count": 0, "new_anchor_count": 0}

        code = extract_code(response)
        success, _ = execute_code(code, work_dir)
        passed = success and os.path.exists(output_path) and evaluate_task(golden_path, output_path, meta["answer_position"])

        old_count = sum(1 for a in manifest.get("old_anchors", []) if len(a) > 2 and a in code)
        new_count = sum(1 for a in manifest.get("new_anchors", []) if len(a) > 2 and a in code)

        return {"task_id": tid, "condition": condition, "patch_type": patch_type,
                "pass": passed, "code": code[:1500], "old_anchor_count": old_count, "new_anchor_count": new_count,
                "model": LLM_MODEL}


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--workers", type=int, default=128)
    parser.add_argument("--repeats", type=int, default=2)
    args = parser.parse_args()

    with open(os.path.join(OUTPUT_DIR, "successful_codes.json")) as f:
        codes = json.load(f)

    with open("data/perturbed_v3/all_manifests.json") as f:
        v3_manifests = json.load(f)

    # Generate perturbed versions + patches
    print(f"Generating perturbed for {len(codes)} tasks...", flush=True)
    manifests = {}
    patches = {}
    for tid in codes:
        m = v3_manifests.get(tid, {})
        col_renames = m.get("spec", {}).get("rename", {}).get("column_rename_map", {})
        sheet_renames = m.get("spec", {}).get("rename", {}).get("sheet_rename_map", {})
        manifest = generate_perturbed(tid, col_renames, sheet_renames)
        manifests[tid] = manifest

        old_anchors = manifest["old_anchors"]
        hard_patch = f"""## Prior Working Solution

Below is a prior working solution for the original version of this task. Reuse it as much as possible, adapting only when necessary.

```python
{codes[tid]}
```"""
        soft_patch = f"""## Prior Solution Note

I previously solved this task. Key details:
- Critical anchors used: {', '.join(f'"{a}"' for a in old_anchors)}
- I used openpyxl to load the workbook, found target columns by exact header name match
- Header row was row 1, data started from row 2
"""
        patches[tid] = {"no_patch": "", "soft_patch": soft_patch, "hard_patch": hard_patch}

    # Build jobs
    jobs = []
    for tid in codes:
        for cond in ["original", "perturbed"]:
            for pt in ["no_patch", "soft_patch", "hard_patch"]:
                for r in range(args.repeats):
                    jobs.append((tid, cond, pt, patches[tid][pt], manifests[tid], r))

    print(f"Running {len(jobs)} jobs with {args.workers} workers...", flush=True)
    results = []
    lock = threading.Lock()
    done = [0]

    def run_job(job):
        tid, cond, pt, patch, manifest, rep = job
        result = run_one(tid, cond, pt, patch, manifest)
        result["repeat"] = rep
        with lock:
            done[0] += 1
            if done[0] % 50 == 0:
                print(f"  [{done[0]}/{len(jobs)}]", flush=True)
        return result

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(run_job, j): j for j in jobs}
        for future in as_completed(futures):
            try:
                results.append(future.result())
            except Exception as e:
                print(f"ERROR: {e}")

    # Save + analyze
    os.makedirs(RESULTS_DIR, exist_ok=True)
    model_short = LLM_MODEL.split('/')[-1].replace(':', '_').replace('.', '_')[:30]
    out_path = os.path.join(RESULTS_DIR, f"positive_control_full_{model_short}.json")
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"Saved to {out_path}")

    print(f"\n{'='*60}\nRESULTS ({len(codes)} tasks)\n{'='*60}")
    for pt in ["no_patch", "soft_patch", "hard_patch"]:
        for cond in ["original", "perturbed"]:
            s = [r for r in results if r["patch_type"] == pt and r["condition"] == cond]
            if s:
                pr = sum(r["pass"] for r in s) / len(s)
                print(f"  {pt:12s} | {cond:10s} | {pr:.1%} ({sum(r['pass'] for r in s)}/{len(s)})")

    print(f"\n{'='*60}\nTRANSFER LOSS\n{'='*60}")
    no_o = [r for r in results if r["patch_type"] == "no_patch" and r["condition"] == "original"]
    no_p = [r for r in results if r["patch_type"] == "no_patch" and r["condition"] == "perturbed"]
    no_or = sum(r["pass"] for r in no_o) / len(no_o) if no_o else 0
    no_pr = sum(r["pass"] for r in no_p) / len(no_p) if no_p else 0
    for pt in ["no_patch", "soft_patch", "hard_patch"]:
        o = [r for r in results if r["patch_type"] == pt and r["condition"] == "original"]
        p = [r for r in results if r["patch_type"] == pt and r["condition"] == "perturbed"]
        if o and p:
            og = sum(r["pass"] for r in o) / len(o)
            pg = sum(r["pass"] for r in p) / len(p)
            gain_o = og - no_or
            gain_p = pg - no_pr
            tl = gain_o - gain_p
            print(f"  {pt:12s} | Gap={og-pg:+.1%} | Gain_orig={gain_o:+.1%} | Gain_pert={gain_p:+.1%} | TL={tl:+.1%}")

    print(f"\n{'='*60}\nOLD ANCHOR USAGE\n{'='*60}")
    for pt in ["no_patch", "soft_patch", "hard_patch"]:
        for cond in ["original", "perturbed"]:
            s = [r for r in results if r["patch_type"] == pt and r["condition"] == cond]
            if s:
                ou = sum(1 for r in s if r.get("old_anchor_count", 0) > 0) / len(s)
                nu = sum(1 for r in s if r.get("new_anchor_count", 0) > 0) / len(s)
                print(f"  {pt:12s} | {cond:10s} | old={ou:.0%} new={nu:.0%}")


if __name__ == "__main__":
    main()
