#!/usr/bin/env python3
"""
Anchor-Dependence Positive Control Experiment.

Goal: Prove that hard local patch (code skeleton with old anchors) becomes brittle
when anchors are opaque-renamed and old-anchor distractors are added.

Design:
- 50 anchor-rich tasks
- Perturb: opaque rename of critical anchors + old-anchor distractor column
- 3 patch types: no_patch, soft_patch, hard_patch
- 2 conditions: original, perturbed
- 2 repeats
- Diagnostic: old_anchor_usage_rate in generated code
"""

import json
import os
import sys
import re
import random
import string
import shutil
import tempfile
import threading
import openpyxl
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.path.insert(0, ".")
from scripts.run_pilot_experiment import (
    call_llm, extract_code, execute_code, evaluate_task, gen_spreadsheet_content,
    VERIFIED_DIR,
)
from scripts.perturbation_agent.llm_client import call_llm as call_llm_bedrock

V3_DIR = "data/perturbed_v3"
OUTPUT_DIR = "data/positive_control"
RESULTS_DIR = "scratch/experiment_results"
LLM_MODEL = "us.anthropic.claude-sonnet-4-5-20250929-v1:0"


# ============================================================
# Step 1: Select 50 anchor-rich tasks
# ============================================================

def select_pilot_tasks(n=50):
    """Select n tasks that have clear column/sheet anchors in instruction."""
    with open(os.path.join(V3_DIR, "all_manifests.json")) as f:
        manifests = json.load(f)
    with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
        dataset = json.load(f)
    meta_map = {str(d["id"]): d for d in dataset}

    candidates = []
    for tid, m in manifests.items():
        if m["status"] != "pass" or tid not in meta_map:
            continue
        # Must have lexical rename (means instruction references headers/sheets)
        if "lexical" not in m.get("anchor_types_used", []):
            continue
        # Must have rename spec with column renames
        spec = m.get("spec", {})
        rename = spec.get("rename", {})
        col_renames = rename.get("column_rename_map", {})
        if not col_renames:
            continue
        meta = meta_map[tid]
        candidates.append({
            "id": tid,
            "family": m.get("family", "unknown"),
            "instruction": meta["instruction"],
            "answer_position": meta["answer_position"],
            "original_col_renames": col_renames,
            "original_sheet_renames": rename.get("sheet_rename_map", {}),
        })

    # Sample diverse tasks
    random.seed(42)
    random.shuffle(candidates)
    selected = candidates[:n]
    print(f"Selected {len(selected)} anchor-rich tasks")
    return selected


# ============================================================
# Step 2: Generate opaque rename + old-anchor distractor
# ============================================================

def gen_opaque_name(prefix="COL"):
    """Generate opaque column/sheet name."""
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3))
    return f"{prefix}_{suffix}"


# Cross-domain natural renames: semantically unrelated but natural-sounding column names
CROSS_DOMAIN_NAMES = [
    "Rating", "Score", "Index", "Level", "Rank", "Grade", "Tier",
    "Segment", "Region", "Zone", "Sector", "Division", "Branch",
    "Label", "Tag", "Code", "Marker", "Flag", "Status", "Phase",
    "Metric", "Factor", "Weight", "Ratio", "Margin", "Delta",
    "Batch", "Cycle", "Period", "Interval", "Span", "Duration",
    "Source", "Origin", "Channel", "Medium", "Platform", "Vendor",
    "Priority", "Severity", "Impact", "Frequency", "Volume",
]

CROSS_DOMAIN_SHEET_NAMES = [
    "Inventory", "Metrics", "Catalog", "Registry", "Ledger",
    "Archive", "Records", "Tracker", "Pipeline", "Workflow",
    "Dashboard", "Overview", "Snapshot", "Digest", "Roster",
]


def gen_natural_rename(old_name, used_names):
    """Generate a natural but semantically unrelated name."""
    random.shuffle(CROSS_DOMAIN_NAMES)
    for name in CROSS_DOMAIN_NAMES:
        if name not in used_names and name.lower() != old_name.lower():
            used_names.add(name)
            return name
    # Fallback
    return gen_opaque_name("COL")


def gen_natural_sheet_rename(old_name, used_names):
    """Generate a natural but semantically unrelated sheet name."""
    random.shuffle(CROSS_DOMAIN_SHEET_NAMES)
    for name in CROSS_DOMAIN_SHEET_NAMES:
        if name not in used_names and name.lower() != old_name.lower():
            used_names.add(name)
            return name
    return gen_opaque_name("WS")


def generate_perturbed_task(task):
    """Generate opaque-renamed workbook with old-anchor distractor."""
    tid = task["id"]
    src_dir = os.path.join(VERIFIED_DIR, "spreadsheet", tid)
    out_dir = os.path.join(OUTPUT_DIR, "perturbed", tid)
    os.makedirs(out_dir, exist_ok=True)

    init_src = os.path.join(src_dir, f"1_{tid}_init.xlsx")
    golden_src = os.path.join(src_dir, f"1_{tid}_golden.xlsx")

    # Load original schema
    wb = openpyxl.load_workbook(init_src)
    original_headers = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val:
                headers.append(str(val))
        original_headers[sn] = headers
    wb.close()

    # Identify critical anchors (from the columns that were renamed in v3)
    col_renames = task["original_col_renames"]
    sheet_renames = task["original_sheet_renames"]

    # Generate natural but semantically unrelated names for critical anchors
    used_names = set()
    opaque_col_map = {}
    for old_name in col_renames.keys():
        opaque_col_map[old_name] = gen_natural_rename(old_name, used_names)

    opaque_sheet_map = {}
    for old_name in sheet_renames.keys():
        opaque_sheet_map[old_name] = gen_natural_sheet_rename(old_name, used_names)

    # Pick the first renamed column as the primary distractor target
    primary_distractor_col = list(col_renames.keys())[0] if col_renames else None

    # Apply to init and golden
    for src_path, dst_name in [(init_src, f"1_{tid}_init.xlsx"), (golden_src, f"1_{tid}_golden.xlsx")]:
        wb = openpyxl.load_workbook(src_path)

        # Rename sheets
        for old_sn, new_sn in opaque_sheet_map.items():
            if old_sn in wb.sheetnames:
                wb[old_sn].title = new_sn

        # Rename headers + add distractor
        for sn in wb.sheetnames:
            ws = wb[sn]
            distractor_added = False
            for c in range(1, ws.max_column + 1):
                cell_val = ws.cell(1, c).value
                if cell_val and str(cell_val) in opaque_col_map:
                    ws.cell(1, c).value = opaque_col_map[str(cell_val)]

            # Add old-anchor distractor column (primary only)
            if primary_distractor_col and not distractor_added:
                # Find if this sheet had the primary column
                orig_sn = sn
                for old_s, new_s in opaque_sheet_map.items():
                    if new_s == sn:
                        orig_sn = old_s
                if primary_distractor_col in original_headers.get(orig_sn, []):
                    # Insert distractor right after the renamed column
                    dist_col = ws.max_column + 1
                    ws.cell(1, dist_col).value = primary_distractor_col
                    # Fill with plausible but wrong numeric data
                    for row in range(2, min(ws.max_row + 1, 200)):
                        ws.cell(row, dist_col).value = round(random.uniform(-100, 100), 2)
                    distractor_added = True

        wb.save(os.path.join(out_dir, dst_name))
        wb.close()

    # Rewrite instruction
    instruction = task["instruction"]
    for old, new in opaque_col_map.items():
        instruction = instruction.replace(old, new)
    for old, new in opaque_sheet_map.items():
        instruction = instruction.replace(old, new)

    # Update answer_position
    ap = task["answer_position"]
    for old, new in opaque_sheet_map.items():
        ap = ap.replace(old, new)

    # Save metadata
    meta = {
        "id": tid,
        "instruction": instruction,
        "answer_position": ap,
        "instruction_type": "Cell-Level Manipulation",
    }
    with open(os.path.join(out_dir, "task_meta.json"), "w") as f:
        json.dump(meta, f, indent=2)

    # Save manifest
    manifest = {
        "task_id": tid,
        "opaque_col_map": opaque_col_map,
        "opaque_sheet_map": opaque_sheet_map,
        "primary_distractor": primary_distractor_col,
        "old_anchors": list(opaque_col_map.keys()) + list(opaque_sheet_map.keys()),
        "new_anchors": list(opaque_col_map.values()) + list(opaque_sheet_map.values()),
    }
    with open(os.path.join(out_dir, "manifest.json"), "w") as f:
        json.dump(manifest, f, indent=2)

    return manifest


# ============================================================
# Step 3: Generate patches (template-based, not LLM free-form)
# ============================================================

def generate_soft_patch(task):
    """Generate soft local patch (natural language with old anchors)."""
    tid = task["id"]
    col_renames = task["original_col_renames"]
    sheet_renames = task["original_sheet_renames"]

    # Get original preview
    init_path = os.path.join(VERIFIED_DIR, "spreadsheet", tid, f"1_{tid}_init.xlsx")
    preview = gen_spreadsheet_content(init_path, max_rows=3)

    old_sheets = list(sheet_renames.keys()) or ["Sheet1"]
    old_cols = list(col_renames.keys())

    patch = f"""## Prior Solution Note

I previously solved this task. Key details:
- The data was in sheet "{old_sheets[0]}"
- Critical columns: {', '.join(f'"{c}"' for c in old_cols)}
- I used openpyxl to load the workbook, found the target columns by matching header names
- Header row was row 1, data started from row 2
- I iterated through rows, applied the required logic, and wrote results to the answer position
- Important: I accessed columns by their exact header name string match, not by position
"""
    return patch


def generate_hard_patch(task):
    """Generate hard local patch (code skeleton with hardcoded old anchors)."""
    tid = task["id"]
    col_renames = task["original_col_renames"]
    sheet_renames = task["original_sheet_renames"]

    old_sheets = list(sheet_renames.keys()) or ["Sheet1"]
    old_cols = list(col_renames.keys())
    primary_sheet = old_sheets[0]
    primary_col = old_cols[0] if old_cols else "Data"

    # Build deterministic code skeleton
    col_find_lines = ""
    for col_name in old_cols[:3]:
        var_name = col_name.lower().replace(" ", "_").replace("-", "_")[:10]
        col_find_lines += f'    if ws.cell(1, c).value == "{col_name}":\n'
        col_find_lines += f'        {var_name}_col = c\n'

    patch = f"""## Prior Solution Code

You should base your solution on the prior solution sketch below. Adapt it minimally to produce the final Python solution.

```python
import openpyxl
import shutil

# Copy input to output first
shutil.copy2(spreadsheet_path, output_path)
wb = openpyxl.load_workbook(output_path)
ws = wb["{primary_sheet}"]

# Find critical columns by header name
{primary_col.lower().replace(" ", "_")[:10]}_col = None
for c in range(1, ws.max_column + 1):
{col_find_lines}
# Read data from row 2 onwards
data = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, {primary_col.lower().replace(" ", "_")[:10]}_col).value
    if val is not None:
        data.append(val)

# Apply task logic and write results
# ... (implement the specific computation here)

wb.save(output_path)
```
"""
    return patch


# ============================================================
# Step 4: Run experiment
# ============================================================

def run_one(task_id, condition, patch_type, patch_content, model, manifest=None):
    """Run a single experiment."""
    if condition == "original":
        spread_dir = os.path.join(VERIFIED_DIR, "spreadsheet", str(task_id))
        with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
            dataset = json.load(f)
        meta = next(d for d in dataset if str(d["id"]) == str(task_id))
    else:
        spread_dir = os.path.join(OUTPUT_DIR, "perturbed", str(task_id))
        with open(os.path.join(spread_dir, "task_meta.json")) as f:
            meta = json.load(f)

    input_file = f"1_{task_id}_init.xlsx"
    golden_file = f"1_{task_id}_golden.xlsx"
    input_path = os.path.join(spread_dir, input_file)
    golden_path = os.path.join(spread_dir, golden_file)

    if not os.path.exists(input_path):
        return {"task_id": task_id, "condition": condition, "patch_type": patch_type,
                "pass": False, "error": "input not found", "code": ""}

    with tempfile.TemporaryDirectory() as work_dir:
        shutil.copy2(input_path, os.path.join(work_dir, input_file))
        output_path = os.path.join(work_dir, f"1_{task_id}_output.xlsx")
        preview = gen_spreadsheet_content(os.path.join(work_dir, input_file))

        artifact_section = ""
        if patch_content:
            artifact_section = f"{patch_content}\n\n"

        prompt = f"""You are a spreadsheet expert who can manipulate spreadsheets through Python code.

{artifact_section}You need to solve the given spreadsheet manipulation question:
### instruction
{meta['instruction']}

### spreadsheet_path
{os.path.join(work_dir, input_file)}

### spreadsheet_content
{preview}

### instruction_type
{meta.get('instruction_type', 'Cell-Level Manipulation')}

### answer_position
{meta['answer_position']}

### output_path
{output_path}

You should generate Python code for the final solution of the question."""

        try:
            response = call_llm(prompt, model, None, None)
        except Exception as e:
            return {"task_id": task_id, "condition": condition, "patch_type": patch_type,
                    "pass": False, "error": str(e), "code": ""}

        code = extract_code(response)
        success, exec_output = execute_code(code, work_dir)
        passed = False
        if success and os.path.exists(output_path):
            passed = evaluate_task(golden_path, output_path, meta["answer_position"])

        # Diagnostic: check old anchor usage in code
        old_anchor_used = []
        if manifest:
            for anchor in manifest.get("old_anchors", []):
                if len(anchor) > 2 and anchor in code:
                    old_anchor_used.append(anchor)

        return {
            "task_id": task_id,
            "condition": condition,
            "patch_type": patch_type,
            "pass": passed,
            "code_executed": success,
            "old_anchor_used": old_anchor_used,
            "old_anchor_count": len(old_anchor_used),
            "code": code[:1000],
            "error": exec_output[:200] if not success else None,
        }


# ============================================================
# Main
# ============================================================

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--n_tasks", type=int, default=50)
    parser.add_argument("--workers", type=int, default=128)
    parser.add_argument("--repeats", type=int, default=2)
    parser.add_argument("--dry_run", action="store_true")
    args = parser.parse_args()

    # Step 1: Select tasks
    tasks = select_pilot_tasks(args.n_tasks)

    # Step 2: Generate perturbed workbooks
    print(f"\nGenerating perturbed workbooks...")
    os.makedirs(os.path.join(OUTPUT_DIR, "perturbed"), exist_ok=True)
    manifests = {}
    for task in tasks:
        m = generate_perturbed_task(task)
        manifests[task["id"]] = m
    print(f"  Generated {len(manifests)} perturbed tasks")

    # Step 3: Generate patches
    print(f"\nGenerating patches...")
    patches = {}
    for task in tasks:
        patches[task["id"]] = {
            "no_patch": "",
            "soft_patch": generate_soft_patch(task),
            "hard_patch": generate_hard_patch(task),
        }
    print(f"  Generated patches for {len(patches)} tasks")

    # Step 4: Build job list
    jobs = []
    for task in tasks:
        tid = task["id"]
        for condition in ["original", "perturbed"]:
            for patch_type in ["no_patch", "soft_patch", "hard_patch"]:
                for repeat in range(args.repeats):
                    jobs.append((tid, condition, patch_type, patches[tid][patch_type],
                                 manifests.get(tid), repeat))

    print(f"\nTotal jobs: {len(jobs)}")
    if args.dry_run:
        print("Dry run — not executing")
        return

    # Step 5: Run
    results = []
    lock = threading.Lock()
    done = [0]
    total = len(jobs)

    def run_job(job):
        tid, condition, patch_type, patch_content, manifest, repeat = job
        result = run_one(tid, condition, patch_type, patch_content, LLM_MODEL, manifest)
        result["repeat"] = repeat
        with lock:
            done[0] += 1
            s = "✅" if result["pass"] else "❌"
            if done[0] % 20 == 0:
                print(f"[{done[0]}/{total}] {s} {tid}|{condition}|{patch_type}", flush=True)
        return result

    print(f"Running {total} jobs with {args.workers} workers...")
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(run_job, j): j for j in jobs}
        for future in as_completed(futures):
            try:
                results.append(future.result())
            except Exception as e:
                job = futures[future]
                print(f"ERROR: {job[0]} {e}")

    # Save results
    os.makedirs(RESULTS_DIR, exist_ok=True)
    results_path = os.path.join(RESULTS_DIR, "positive_control_results.json")
    with open(results_path, "w") as f:
        json.dump(results, f, indent=2)

    # Step 6: Analysis
    print(f"\n{'='*60}")
    print("RESULTS")
    print(f"{'='*60}")
    for patch_type in ["no_patch", "soft_patch", "hard_patch"]:
        for condition in ["original", "perturbed"]:
            subset = [r for r in results if r["patch_type"] == patch_type and r["condition"] == condition]
            if subset:
                pr = sum(r["pass"] for r in subset) / len(subset)
                print(f"  {patch_type:12s} | {condition:10s} | {pr:.1%} ({sum(r['pass'] for r in subset)}/{len(subset)})")

    print(f"\n{'='*60}")
    print("PATCH GAP")
    print(f"{'='*60}")
    for patch_type in ["no_patch", "soft_patch", "hard_patch"]:
        orig = [r for r in results if r["patch_type"] == patch_type and r["condition"] == "original"]
        pert = [r for r in results if r["patch_type"] == patch_type and r["condition"] == "perturbed"]
        if orig and pert:
            og = sum(r["pass"] for r in orig) / len(orig)
            pg = sum(r["pass"] for r in pert) / len(pert)
            print(f"  {patch_type:12s} | Gap = {og-pg:+.1%}  ({og:.1%} -> {pg:.1%})")

    print(f"\n{'='*60}")
    print("OLD ANCHOR USAGE (perturbed condition only)")
    print(f"{'='*60}")
    for patch_type in ["no_patch", "soft_patch", "hard_patch"]:
        subset = [r for r in results if r["patch_type"] == patch_type and r["condition"] == "perturbed"]
        if subset:
            usage = sum(1 for r in subset if r.get("old_anchor_count", 0) > 0) / len(subset)
            print(f"  {patch_type:12s} | old_anchor_usage = {usage:.0%}")

    print(f"\nResults saved to {results_path}")


if __name__ == "__main__":
    main()
