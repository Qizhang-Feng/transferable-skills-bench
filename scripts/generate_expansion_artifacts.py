#!/usr/bin/env python3
"""
Generate per-task local_patch artifacts for expansion tasks.
Each task's local_patch is based on its paired task's specific details.
"""

import json
import os
import openpyxl

VERIFIED_DIR = "vendor/SpreadsheetBench/data/spreadsheetbench_verified_400"


def get_task_details(task_id):
    """Get instruction, sheet names, headers, and answer_position for a task."""
    with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
        dataset = json.load(f)
    meta = next((d for d in dataset if str(d["id"]) == str(task_id)), None)
    if not meta:
        return None

    spread_dir = os.path.join(VERIFIED_DIR, "spreadsheet", str(task_id))
    init_file = os.path.join(spread_dir, f"1_{task_id}_init.xlsx")

    sheets_info = []
    try:
        wb = openpyxl.load_workbook(init_file)
        for sn in wb.sheetnames:
            ws = wb[sn]
            headers = []
            for cell in ws[1]:
                if cell.value is not None:
                    headers.append({"letter": cell.column_letter, "name": str(cell.value)})
            sheets_info.append({
                "name": sn,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "headers": headers,
            })
        wb.close()
    except Exception:
        pass

    return {
        "task_id": task_id,
        "instruction": meta["instruction"],
        "instruction_type": meta["instruction_type"],
        "answer_position": meta["answer_position"],
        "sheets": sheets_info,
    }


def generate_local_patch(source_task_details):
    """Generate a local_patch artifact based on a source task's specific details."""
    tid = source_task_details["task_id"]
    inst = source_task_details["instruction"]
    sheets = source_task_details["sheets"]
    ap = source_task_details["answer_position"]

    # Build sheet-specific details
    sheet_details = []
    for sh in sheets:
        headers_str = ", ".join(
            f"'{h['name']}' (column {h['letter']})" for h in sh["headers"][:8]
        )
        sheet_details.append(
            f"- Sheet '{sh['name']}' has {sh['rows']} rows and {sh['cols']} columns. "
            f"Headers: {headers_str}."
        )

    sheets_section = "\n".join(sheet_details)

    # Extract key instruction details (first 300 chars)
    inst_summary = inst[:300]
    if len(inst) > 300:
        inst_summary += "..."

    patch = f"""## Local Patch — Based on Task {tid}

### Context
This patch is based on solving a spreadsheet task with the following setup:

{sheets_section}

The answer should be written to: {ap}

### Task Summary
{inst_summary}

### Steps
1. Open the workbook. The data is in the '{sheets[0]['name']}' sheet.
"""

    # Add header-specific steps
    if sheets[0]["headers"]:
        for h in sheets[0]["headers"][:5]:
            patch += f"2. The '{h['name']}' data is in column {h['letter']}.\n"
            break  # Just show the first one as example

    patch += f"""3. Read the data starting from row 2 (row 1 is headers).
4. Implement the logic described in the instruction.
5. Write results to the target cells at {ap}.
6. Save the workbook to the output path.

### Known details
- The main data sheet is '{sheets[0]['name']}'.
"""

    if len(sheets) > 1:
        patch += f"- There are {len(sheets)} sheets total: {', '.join(sh['name'] for sh in sheets)}.\n"

    for h in sheets[0]["headers"][:6]:
        patch += f"- Column {h['letter']} contains '{h['name']}'.\n"

    patch += f"- The answer position is {ap}.\n"

    return patch


def main():
    with open("scratch/expansion-pairings.json") as f:
        pairings = json.load(f)

    with open("scratch/task-audit-60-v1.json") as f:
        audit = json.load(f)
    audit_map = {str(t["task_id"]): t for t in audit["tasks"]}

    for target_tid, source_tid in pairings.items():
        # Get source task details (the task the patch is based on)
        source_details = get_task_details(source_tid)
        if not source_details:
            print(f"WARNING: Could not get details for source task {source_tid}")
            continue

        # Generate local_patch
        patch_content = generate_local_patch(source_details)

        # Save
        out_dir = os.path.join("scratch/artifact_baselines/per_task", str(target_tid))
        os.makedirs(out_dir, exist_ok=True)

        patch_path = os.path.join(out_dir, "local_patch.md")
        with open(patch_path, "w") as f:
            f.write(patch_content)

        print(f"  {target_tid:12s} <- patch from {source_tid:12s} ({audit_map.get(target_tid, {}).get('workflow_family_v1', '?')})")

    print(f"\nGenerated {len(pairings)} local_patch artifacts")


if __name__ == "__main__":
    main()
