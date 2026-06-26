#!/usr/bin/env python3
"""
Extract detailed schema info for candidate tasks.
For each task: sheet names, headers, instruction text, column names mentioned in instruction.
Output: JSON file ready for perturbation spec generation.
"""

import json
import os
import re
import openpyxl

VERIFIED_DIR = "vendor/SpreadsheetBench/data/spreadsheetbench_verified_400"

CANDIDATES = [
    # conditional_partition
    "341-40", "24-23", "73-45",
    # formula_writeback easy
    "1818", "31915", "49801", "36191", "49857", "58723", "57558",
    # grouped_aggregation
    "52575", "203-15", "395-36",
    # lookup_match_merge
    "146-49",
    # structural_edit easy
    "374-31", "367-23", "486-17",
    # structural_edit medium
    "577-40", "147-48", "493-5", "79-7", "488-14",
]


def get_headers(ws, max_scan_rows=5):
    """Get header values from first non-empty row."""
    for row_idx in range(1, max_scan_rows + 1):
        vals = []
        for cell in ws[row_idx]:
            if cell.value is not None:
                vals.append({"col_letter": cell.column_letter, "value": str(cell.value)[:60]})
        if vals:
            return {"row": row_idx, "headers": vals}
    return {"row": None, "headers": []}


def count_formulas(ws):
    count = 0
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count


def find_names_in_instruction(instruction, sheet_names, header_values):
    """Find which sheet names and column headers are mentioned in the instruction."""
    mentioned_sheets = []
    for sn in sheet_names:
        if len(sn) <= 2:
            # Short names: require quotes or exact word boundary
            patterns = [f"'{sn}'", f'"{sn}"', f"'{sn}'"]
            for p in patterns:
                if p in instruction:
                    mentioned_sheets.append(sn)
                    break
        else:
            if sn.lower() in instruction.lower():
                mentioned_sheets.append(sn)

    mentioned_headers = []
    for hv in header_values:
        if len(hv) <= 2:
            continue  # Skip single-letter headers
        # Case-insensitive search with word boundary
        pattern = re.compile(r'\b' + re.escape(hv) + r'\b', re.IGNORECASE)
        if pattern.search(instruction):
            mentioned_headers.append(hv)

    return list(set(mentioned_sheets)), list(set(mentioned_headers))


def main():
    with open(os.path.join(VERIFIED_DIR, "dataset.json")) as f:
        dataset = json.load(f)
    meta_map = {str(d["id"]): d for d in dataset}

    with open("scratch/task-audit-60-v1.json") as f:
        audit = json.load(f)
    audit_map = {str(t["task_id"]): t for t in audit["tasks"]}

    results = []
    for tid in CANDIDATES:
        meta = meta_map.get(tid, {})
        audit_info = audit_map.get(tid, {})
        spread_dir = os.path.join(VERIFIED_DIR, "spreadsheet", tid)
        init_file = os.path.join(spread_dir, f"1_{tid}_init.xlsx")

        try:
            wb = openpyxl.load_workbook(init_file)
        except Exception as e:
            results.append({"task_id": tid, "error": str(e)[:100]})
            continue

        sheets_info = []
        all_header_values = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            header_info = get_headers(ws)
            formulas = count_formulas(ws)
            merged = len(ws.merged_cells.ranges)

            header_vals = [h["value"] for h in header_info["headers"]]
            all_header_values.extend(header_vals)

            sheets_info.append({
                "name": sn,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "header_row": header_info["row"],
                "headers": header_info["headers"],
                "formula_count": formulas,
                "merged_cell_count": merged,
            })

        instruction = meta.get("instruction", "")
        mentioned_sheets, mentioned_headers = find_names_in_instruction(
            instruction, wb.sheetnames, all_header_values
        )

        results.append({
            "task_id": tid,
            "family": audit_info.get("workflow_family_v1", ""),
            "perturbability": audit_info.get("perturbability", ""),
            "instruction": instruction,
            "instruction_type": meta.get("instruction_type", ""),
            "answer_position": meta.get("answer_position", ""),
            "sheets": sheets_info,
            "sheet_names": wb.sheetnames,
            "mentioned_sheets": mentioned_sheets,
            "mentioned_headers": mentioned_headers,
        })
        wb.close()

    with open("scratch/expansion-task-schemas.json", "w") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"Extracted schemas for {len(results)} tasks -> scratch/expansion-task-schemas.json")

    # Summary
    for r in results:
        if "error" in r:
            print(f"  {r['task_id']:12s} ERROR: {r['error']}")
            continue
        ms = ", ".join(r["mentioned_sheets"][:3]) or "(none)"
        mh = ", ".join(r["mentioned_headers"][:5]) or "(none)"
        print(f"  {r['task_id']:12s} | {r['family']:25s} | sheets: {ms:30s} | headers: {mh}")


if __name__ == "__main__":
    main()
